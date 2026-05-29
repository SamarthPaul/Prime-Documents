"""
Merge a Playwright JSON run into the QA execution workbook.

    python _merge_results.py [path/to/results.json]

- Reads the Playwright JSON reporter output (default:
  tests/playwright/test-results/results.json).
- For every test whose title contains a PR-TC-* id, records the outcome.
- Updates the 'Test Cases' sheet: Status (Pass/Fail/Blocked/Flaky), Actual Result
  (error snippet or 'OK'), and Severity for failures (from Priority).
- Rebuilds the 'Bugs Only' sheet from the failures.

Outcome mapping uses Playwright's test-level status so that test.fail/test.fixme
placeholders do NOT show up as bugs:
    expected   -> Pass     (passed, or failed-as-expected placeholder)
    unexpected -> Fail     (a real failure = a bug)
    flaky      -> Flaky    (passed on retry — investigate)
    skipped    -> Blocked  (precondition not met, e.g. missing admin creds)
"""
import json, os, re, sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(__file__)
XLSX = os.path.join(HERE, 'test-cases.xlsx')
RESULTS = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
    HERE, 'tests', 'playwright', 'test-results', 'results.json')

ID_RE = re.compile(r'PR-TC-[A-Z]+(?:-[A-Z]+)*-\d+[a-z]?')

# Redact entrepreneur record names that leak into "visible on <X> (<record>):"
# messages — the workbook ships to a PUBLIC repo, so no real PII.
_PII_RE = re.compile(r'(visible on)\b.*?(?=:|\{|$)', re.IGNORECASE)
def redact_pii(s):
    if not isinstance(s, str) or 'visible on' not in s:
        return s
    # drop every parenthetical (record names / labels) in the leak phrase
    return re.sub(r'\([^)]*\)', '', s).replace('  ', ' ').replace(' :', ':')
SEVERITY_BY_PRI = {'P0': 'Critical', 'P1': 'High', 'P2': 'Medium'}

STATUS_MAP = {'expected': 'Pass', 'unexpected': 'Fail', 'flaky': 'Flaky', 'skipped': 'Blocked'}

# ---- walk the Playwright JSON tree -----------------------------------------
def walk_specs(node, out):
    for spec in node.get('specs', []):
        for t in spec.get('tests', []):
            ids = ID_RE.findall(spec.get('title', ''))
            if not ids:
                continue
            res = (t.get('results') or [{}])[-1]
            err = ''
            if res.get('error'):
                err = res['error'].get('message', '')
            elif res.get('errors'):
                err = res['errors'][0].get('message', '')
            shot = ''
            for a in res.get('attachments', []):
                if a.get('name') == 'screenshot' and a.get('path'):
                    shot = a['path']; break
            out.append({
                'id': ids[0],
                'title': spec.get('title', ''),
                'project': t.get('projectName', ''),
                'status': t.get('status', 'unknown'),
                'error': re.sub(r'\x1b\[[0-9;]*m', '', err).strip()[:500],  # strip ANSI
                'screenshot': shot,
            })
    for child in node.get('suites', []):
        walk_specs(child, out)

if not os.path.exists(RESULTS):
    sys.exit(f'No results file at {RESULTS} — run the suite first.')

with open(RESULTS, encoding='utf-8') as f:
    data = json.load(f)

records = []
for s in data.get('suites', []):
    walk_specs(s, records)

# Aggregate per test-case id across projects. Fail wins, then Flaky, Blocked, Pass.
RANK = {'Fail': 3, 'Flaky': 2, 'Blocked': 1, 'Pass': 0}
agg = {}
for r in records:
    st = STATUS_MAP.get(r['status'], 'Fail')
    cur = agg.get(r['id'])
    cand = {'status': st, 'projects': {r['project']}, 'error': r['error'], 'screenshot': r['screenshot'], 'title': r['title']}
    if not cur:
        agg[r['id']] = cand
    else:
        cur['projects'].add(r['project'])
        if RANK[st] > RANK[cur['status']]:
            cur['status'] = st
            if r['error']: cur['error'] = r['error']
            if r['screenshot']: cur['screenshot'] = r['screenshot']

print(f'Parsed {len(records)} test runs → {len(agg)} distinct test-case ids')

# ---- fold in SEMI verdicts (MCP visual pass) --------------------------------
# _semi_verdicts.json: { "PR-TC-XC-052": {"status":"Fail","actual":"...","screenshot":"..."}, ... }
# AUTO (Playwright) takes precedence if an id appears in both.
SEMI = os.path.join(HERE, '_semi_verdicts.json')
if os.path.exists(SEMI):
    with open(SEMI, encoding='utf-8') as f:
        semi = json.load(f)
    added = 0
    for sid, v in semi.items():
        # A SEMI verdict overrides only a Blocked/absent AUTO entry — a real
        # AUTO Pass/Fail/Flaky always wins over a manual verdict.
        if sid in agg and agg[sid]['status'] != 'Blocked':
            continue
        agg[sid] = {'status': v.get('status', 'Blocked'), 'projects': {'semi'},
                    'error': v.get('actual', ''), 'screenshot': v.get('screenshot', ''), 'title': v.get('title', sid)}
        added += 1
    print(f'Folded in {added} SEMI verdicts from _semi_verdicts.json')

# ---- update the workbook ----------------------------------------------------
wb = load_workbook(XLSX)
tc = wb['Test Cases']
hdr = {c.value: i + 1 for i, c in enumerate(tc[1])}
col_id, col_pri = hdr['ID'], hdr['Priority']
col_status, col_actual, col_sev = hdr['Status'], hdr['Actual Result'], hdr['Severity']

# index rows by id, and capture fields needed for the Bugs sheet
row_by_id = {}
for r in range(2, tc.max_row + 1):
    rid = tc.cell(r, col_id).value
    if rid:
        row_by_id[rid.strip()] = r

status_fill = {
    'Pass':    PatternFill('solid', fgColor='D9E8E5'),
    'Fail':    PatternFill('solid', fgColor='F3C9B6'),
    'Flaky':   PatternFill('solid', fgColor='F5E2D6'),
    'Blocked': PatternFill('solid', fgColor='EDE6D2'),
}
matched, unmatched = 0, []
for tcid, a in agg.items():
    r = row_by_id.get(tcid)
    if not r:
        unmatched.append(tcid); continue
    matched += 1
    tc.cell(r, col_status, a['status'])
    tc.cell(r, col_status).fill = status_fill.get(a['status'])
    tc.cell(r, col_actual, redact_pii(a['error']) if a['status'] in ('Fail', 'Flaky') else 'OK')
    if a['status'] == 'Fail':
        pri = (tc.cell(r, col_pri).value or '')
        pcode = next((k for k in SEVERITY_BY_PRI if k in pri), 'P1')
        tc.cell(r, col_sev, SEVERITY_BY_PRI[pcode])

print(f'  matched {matched} ids into Test Cases sheet'
      + (f'; {len(unmatched)} ids not found: {unmatched}' if unmatched else ''))

# ---- rebuild Bugs Only ------------------------------------------------------
bugs = wb['Bugs Only']
# clear old data rows (keep note row 1 + header row 2)
if bugs.max_row > 2:
    bugs.delete_rows(3, bugs.max_row - 2)

body_font = Font(name='Georgia', size=10)
thin = Side(border_style='thin', color='E8DFC9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical='top', wrap_text=True)

def cellval(r, name):
    return tc.cell(r, hdr[name]).value or ''

bug_n = 0
for tcid, a in sorted(agg.items()):
    if a['status'] != 'Fail':
        continue
    r = row_by_id.get(tcid)
    if r:
        pri = cellval(r, 'Priority')
        pcode = next((k for k in SEVERITY_BY_PRI if k in str(pri)), 'P1')
        module, title = cellval(r, 'Module'), cellval(r, 'Title')
        steps, expected, brd = cellval(r, 'Steps'), cellval(r, 'Expected Result'), cellval(r, 'BRD Ref')
    else:
        # Failing check not yet in the master (e.g. the UI-LEAK class added from a
        # UAT observation). Still report it so nothing is lost; add to master later.
        pri = ''
        module = '(not in master — add as test case)'
        title = a.get('title', tcid)
        steps = expected = brd = ''
        pcode = 'High' if re.search(r'LEAK|SEC', tcid) else 'Medium'
        pcode = {'High': 'P0', 'Medium': 'P1'}.get(pcode, 'P1')  # normalise for severity map below
    bug_n += 1
    bugs.append([
        f'BUG-{bug_n:03d}', tcid, module, title,
        SEVERITY_BY_PRI[pcode], pri, steps, expected,
        redact_pii(a['error']), a['screenshot'], brd, 'Open', '', '',
    ])
    rn = bugs.max_row
    for col in range(1, 15):
        c = bugs.cell(rn, col); c.font = body_font; c.border = border; c.alignment = wrap

wb.save(XLSX)

# summary
from collections import Counter
counts = Counter(a['status'] for a in agg.values())
print(f'Wrote {XLSX}')
print(f'  Status counts: {dict(counts)}')
print(f'  Bugs Only rows: {bug_n}')
