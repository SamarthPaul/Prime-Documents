# -*- coding: utf-8 -*-
"""Populate the Prime Rural Bug Tracker (Internal QA sheet) from row 31 with
bugs found by Samarth while testing staging (3 Jun 2026).

Rules honoured:
- Write into existing 'Internal QA' sheet starting at row 31.
- Match column layout (A..V) and existing conventions (rows 8-29 by Samarth).
- Choose values that match the in-sheet dropdowns / used vocab.
- DO NOT merge cells. DO NOT add colour. DO NOT invent extra info.
- Copy the (colour-less) cell style from row 29 so borders/wrap match.
"""
import copy
import datetime
import shutil
import os
import openpyxl

# Canonical tracker now lives in generated/ only (from-client/ copy was deleted
# 4 Jun 2026 per user). The script loads and writes that same file in place;
# rows 31+ are overwritten on each run, so re-running is idempotent.
SRC = r"generated/Prime_Rural_Bug_Tracker_updated.xlsx"
GEN = r"generated/Prime_Rural_Bug_Tracker_updated.xlsx"

RAISED_BY = "Samarth Paul"
RAISE_DATE = datetime.datetime(2026, 6, 3)

# Column order A..V
# A ID | B Environment | C Module | D Type of Bug | E Type of Testing |
# F User id | G Bug Location | H Interface | I Short Desc | J Long Desc |
# K Intended Flow | L Assigned | M Priority | N Link | O Raised By |
# P Raising Date | Q Dev Status | R Dev Comment | S QA Status | T QA Remark |
# U Overall Status | V Closing Date

# Each bug: (Module C, TypeOfBug D, Testing E, UserId F, Location G, Interface H,
#            Short I, Long J, Intended K, Priority M)
BUGS = [
    ("Landscape", "Improvement", "Functional Testing", "Admin",
     "Landscape → Overview Cards", "Web",
     "Introduce MET / Incubated / Supported cards on Landscape",
     "On the Landscape view the total entrepreneurs count should be shown as \"MET Entrepreneurs\". "
     "Within MET, two sub-cards — Incubated and Supported — should be displayed, driven by the "
     "final selection field in the EP Profile form.",
     "Landscape shows a MET Entrepreneurs card (all saved profiles) with two sub-cards (Incubated, "
     "Supported) populated from the EP Profile final selection field.",
     "P2"),

    ("Entrepreneur Profile → Location and Contact", "Improvement", "Functional Testing", "Admin",
     "EP Profile → Location and Contact", "Web",
     "Allow adding a village to Village Master from Location & Contact (Core Team IT role)",
     "In the Location and Contact section of the EP Profile form there is no option to add a new "
     "village to the Village Master. Provide an add-village option for the Core Team IT user role.",
     "Core Team IT users can add a new village to the Village Master directly from the Location and "
     "Contact section of the EP Profile form.",
     "P2"),

    ("Masters → Village Master", "Improvement", "Functional Testing", "Admin",
     "Village Master → Add Village", "Web",
     "Add 'GS Circle Name' field in Village Master",
     "The Village Master is missing a GS Circle Name field. Add a GS Circle Name field whose dropdown "
     "values come from the master shared by the client.",
     "Village Master has a GS Circle Name field as a dropdown sourced from the client-shared master.",
     "P2"),

    ("Village Visit Form → Identity Tab", "UI / Cosmetic", "UI testing", "Admin",
     "Village Visit Form → Identity Tab", "Web",
     "Extend width of 'About Village' text box to section width",
     "In the Identity tab of the Village Visit form the About Village text box is narrower than the "
     "section. Extend its width to match the full section width.",
     "About Village text box spans the full width of its section in the Identity tab.",
     "P3"),

    ("Village Visit Form → Demographics Tab", "UI / Cosmetic", "UI testing", "Admin",
     "Village Visit Form → Demographics Tab", "Web",
     "Extend width of 'Communities' text box to section width",
     "In the Demographics tab of the Village Visit form the Communities text box is narrower than the "
     "section. Extend its width to match the full section width.",
     "Communities text box spans the full width of its section in the Demographics tab.",
     "P3"),

    ("Map", "Improvement", "Functional Testing", "Admin",
     "Map Component → Geo Coordinates", "Both",
     "Allow setting geo coordinates by moving pin on map when offline",
     "When there is no internet connectivity, the user should still be able to set the geo "
     "coordinates by dragging/moving the pin on the map.",
     "With no internet, the user can set geo coordinates by moving the pin on the map.",
     "P2"),

    ("Entrepreneur Profile → Photo Gallery", "Improvement", "Functional Testing", "Admin",
     "EP Profile → Profile Picture & Photo Gallery", "Both",
     "Add take-picture / upload option for profile picture and photo gallery",
     "Provide a take picture (camera capture) or upload option for the profile picture field in the "
     "EP Profile and for the Photo Gallery.",
     "Profile picture and Photo Gallery both allow capturing a new photo or uploading an existing one.",
     "P2"),

    ("Entrepreneur Profile → About Enterprise", "Improvement", "Functional Testing", "Admin",
     "EP Profile → About Enterprise", "Web",
     "Add Sector multiselect and Business Basket multiselect fields",
     "In the About Enterprise section the Sector field and the Business Basket field should be "
     "multiselect fields.",
     "About Enterprise section has Sector and Business Basket as multiselect fields.",
     "P2"),

    ("Entrepreneur Profile → Profile Tab", "Improvement", "Functional Testing", "Admin",
     "EP Profile → Profile Tab", "Web",
     "Fix Profile tab section sequence, relabel 'Initial Intervention Analysis', allow multi image upload",
     "The Profile tab sections should follow this sequence: Project Selection, then About Entrepreneur, "
     "then About Enterprise, then Products, then Initial Need Analysis (rename the existing "
     "'Initial Intervention Analysis' label to 'Initial Need Analysis'), then Photo Gallery. The Photo "
     "Gallery section should allow multiple image upload.",
     "Profile tab sections appear in order: Project Selection → About Entrepreneur → About "
     "Enterprise → Products → Initial Need Analysis → Photo Gallery (multiple uploads "
     "allowed); the section is labelled 'Initial Need Analysis'.",
     "P2"),

    ("Entrepreneur Profile → Initial Need Analysis", "Improvement", "Functional Testing", "Admin",
     "EP Profile → Initial Need Analysis Section", "Web",
     "Add 'If any other' checkbox with comment in Initial Need Analysis",
     "The Initial Need Analysis section (in the Profile tab of the EP Profile form) should include an "
     "'If any other' checkbox along with a comment field.",
     "Initial Need Analysis section has an 'If any other' checkbox with an accompanying comment field.",
     "P3"),

    ("Entrepreneur Profile", "Validation Issue", "Functional Testing", "Admin",
     "EP Profile → Save / Form Validation", "Web",
     "On save, highlight the exact error (missing mandatory field, site, or document number validation)",
     "When saving the EP Profile form, the system should highlight the exact error — e.g. which "
     "mandatory field is missing, if the site is not added correctly, or if a document number "
     "validation is failing — instead of a generic error.",
     "On save, the form points to the specific failing field/condition (mandatory field, site, or "
     "document number) with a clear message.",
     "P1"),

    ("Entrepreneur Profile → Location and Contact", "Bug", "Functional Testing", "Admin",
     "EP Profile → Location and Contact → Mobile No", "Web",
     "Extra zero added to Mobile No on save",
     "On saving the EP Profile, an extra zero appears in the Mobile No field — the value stored differs "
     "from what the user entered.",
     "Mobile No is stored exactly as entered, with no extra digit appended on save.",
     "P1"),

    ("Entrepreneur Profile → Location and Contact", "Validation Issue", "Functional Testing", "Admin",
     "EP Profile → Location and Contact → Mobile No", "Web",
     "Mobile No missing 10-digit / starts-with-6-7-8-9 validation",
     "The Mobile No field does not validate input. It should require a 10-digit number that starts with "
     "6, 7, 8 or 9, and show an error on wrong input when saving.",
     "Mobile No is validated to be exactly 10 digits starting with 6/7/8/9; invalid input is blocked on "
     "save with a clear error.",
     "P1"),

    ("Log Book", "Improvement", "Functional Testing", "Admin",
     "Log Book → Photo Capture", "Both",
     "Log book photos: add capture + upload, and show GPS coordinates and date/time",
     "In the Log Book, open the camera capture feature and also provide an upload option. For photos "
     "captured, the GPS coordinates and the date and time should be shown.",
     "Log Book supports capturing or uploading photos; each photo shows its GPS coordinates and "
     "capture date/time.",
     "P2"),

    ("Prioritisation Tab", "Improvement", "Functional Testing", "Admin",
     "Prioritisation Tab", "Web",
     "Remove 'Add Row' in Prioritisation tab",
     "The Prioritisation tab currently allows adding rows. Remove the Add Row option from the "
     "Prioritisation tab.",
     "Prioritisation tab has no Add Row option.",
     "P3"),

    ("Prioritisation Tab", "Validation Issue", "Functional Testing", "Admin",
     "Prioritisation Tab", "Web",
     "Prevent duplicate prioritisation numbers across frameworks",
     "No two frameworks can have the same prioritisation number. Give a relevant error message, or do "
     "not allow the input, if a number is repeated.",
     "Prioritisation numbers are unique across frameworks; repeating a number is blocked or shows a "
     "clear error.",
     "P2"),

    ("Overview Page", "Bug", "Functional Testing", "Admin",
     "Overview Page", "Web",
     "Hide frameworks with prioritisation 0 from the Overview page",
     "Frameworks that have 0 in the Prioritisation tab should not appear (vanish) on the Overview page.",
     "Frameworks with a prioritisation value of 0 are not shown on the Overview page.",
     "P2"),

    ("Overview Page", "UI / Cosmetic", "UI testing", "Admin",
     "Overview Page", "Web",
     "Numbers should be terracotta instead of purple on Overview page",
     "On the Overview page the numbers are shown in purple. They should be shown in terracotta instead.",
     "Numbers on the Overview page are displayed in terracotta.",
     "P3"),

    ("Landscape → Map", "UI / Cosmetic", "UI testing", "Admin",
     "Landscape → Map Header", "Web",
     "Make the Meghalaya map header banner teal on Landscape",
     "On the Landscape view, the map header banner where 'Meghalaya' is written should be teal.",
     "The Landscape map header banner (Meghalaya) is teal.",
     "P3"),

    ("Overview Tab", "Bug", "Functional Testing", "Admin",
     "Overview Tab", "Web",
     "District and Village should populate in Overview tab once added",
     "The District and Village values are not populating in the Overview tab after they have been "
     "added. They should populate once added.",
     "District and Village appear in the Overview tab as soon as they are added.",
     "P2"),

    ("Entrepreneur Profile → Intervention Logs", "Improvement", "Functional Testing", "Admin",
     "EP Profile → Intervention Logs", "Web",
     "Remove 'Add Log' from EP Intervention Logs (read-only section)",
     "The EP Intervention Logs is a read-only section — all logs of that entrepreneur from the "
     "frameworks populate here automatically. Remove the Add Log option.",
     "EP Intervention Logs has no Add Log option; it only shows logs auto-populated from the "
     "entrepreneur's frameworks.",
     "P2"),

    ("All Modules", "Improvement", "UI testing", "Admin",
     "All Modules → Date Fields", "Web",
     "Use DD-MM-YYYY format for all date fields everywhere",
     "All date fields across the application should use the DD-MM-YYYY format.",
     "Every date field is displayed in DD-MM-YYYY format.",
     "P2"),

    ("Entrepreneur Profile → About Entrepreneur", "Validation Issue", "Functional Testing", "Admin",
     "EP Profile → About Entrepreneur → DOB", "Web",
     "DOB field should not allow future dates",
     "The Date of Birth field should not allow a future date to be selected.",
     "The DOB date picker disallows selecting/saving any future date.",
     "P1"),

    ("Entrepreneur Profile → Networking & Mentoring", "Improvement", "Functional Testing", "Admin",
     "EP Profile → Networking & Mentoring Tab", "Web",
     "Allow Fellow to add a mentor in Networking & Mentoring tab (Mentor Master access to field user)",
     "In the Networking and Mentoring tab, give the Fellow an option to add a mentor — i.e. give "
     "the field user access to the Mentor Master only.",
     "Fellows (field users) can add a mentor from the Networking & Mentoring tab, with access limited "
     "to the Mentor Master.",
     "P2"),

    ("Log Book", "Improvement", "Functional Testing", "Admin",
     "Log Book Form", "Web",
     "Add 'Hours' field/column in Log Book",
     "Add an Hours field as a column in the Log Book form so the Fellow can record how many hours were "
     "taken for that activity.",
     "Log Book has an Hours field/column for the Fellow to enter time spent on the activity.",
     "P2"),

    ("Entrepreneur Profile → About Enterprise", "Improvement", "Functional Testing", "Admin",
     "EP Profile → About Enterprise", "Web",
     "Add 'Legal Status' field in About Enterprise",
     "Add a Legal Status field in the About Enterprise section of the Profile tab in the EP Profile form.",
     "About Enterprise section (Profile tab) has a Legal Status field.",
     "P2"),

    ("Entrepreneur Profile", "Validation Issue", "Functional Testing", "Admin",
     "EP Profile → Save / Form Validation", "Web",
     "Saving an empty form should show mandatory-field error for Name and Mobile No",
     "If the user saves an empty form, the system should show an error message that the mandatory "
     "fields — Name and Mobile No — must be filled.",
     "Saving an empty EP Profile shows a clear error that Name and Mobile No are mandatory.",
     "P1"),
]

wb = openpyxl.load_workbook(SRC)
ws = wb["Internal QA"]

# Style template row (colour-less) to copy from
TEMPLATE_ROW = 29
COLS = list(range(1, 23))  # A..V

start_row = 31
for i, bug in enumerate(BUGS):
    r = start_row + i
    module, typ, testing, uid, loc, iface, short, long, intended, prio = bug
    # copy style from template row first (borders, wrap, number formats) - no colour
    for c in COLS:
        src = ws.cell(row=TEMPLATE_ROW, column=c)
        tgt = ws.cell(row=r, column=c)
        tgt._style = copy.copy(src._style)
    ws.cell(row=r, column=2, value="Staging")        # B Environment
    ws.cell(row=r, column=3, value=module)            # C Module
    ws.cell(row=r, column=4, value=typ)               # D Type of Bug
    ws.cell(row=r, column=5, value=testing)           # E Type of Testing
    ws.cell(row=r, column=6, value=uid)               # F User id
    ws.cell(row=r, column=7, value=loc)               # G Bug Location
    ws.cell(row=r, column=8, value=iface)             # H Interface
    ws.cell(row=r, column=9, value=short)             # I Short Desc
    ws.cell(row=r, column=10, value=long)             # J Long Desc
    ws.cell(row=r, column=11, value=intended)         # K Intended Flow
    ws.cell(row=r, column=13, value=prio)             # M Priority
    ws.cell(row=r, column=15, value=RAISED_BY)        # O Raised By
    ws.cell(row=r, column=16, value=RAISE_DATE)       # P Raising Date
    ws.cell(row=r, column=17, value="Open")           # Q Dev Status
    ws.cell(row=r, column=19, value="Open")           # S QA Status
    ws.cell(row=r, column=21, value="Open")           # U Overall Status

os.makedirs("generated", exist_ok=True)
wb.save(GEN)
print("Saved", GEN, "with", len(BUGS), "bugs at rows", start_row, "to", start_row + len(BUGS) - 1)
