"""
Parse generated/test-cases.html and emit generated/test-cases.xlsx — the QA
EXECUTION workbook used to run the UAT pass and hand a bug report to the dev.

Sheets:
  - Summary     : coverage counts incl. AUTO (Playwright) vs SEMI (MCP + manual)
  - Test Cases  : all 319 cases + execution columns (Status / Actual / Severity / Comments)
  - Bugs Only   : failures only — populated by _merge_results.py after a run

Columns (Test Cases):
  ID | Module | Subsection | Title | Type | Exec Mode | Polarity | Priority | Role |
  Preconditions | Steps | Expected Result | BRD Ref | Status | Actual Result | Severity | Comments

Exec Mode split (validate this!):
  AUTO = Functional, Validation, API, Database, Security   -> driven by Playwright
  SEMI = UI / UX, Non-Functional                           -> MCP screenshots + manual observation
"""
import re, os
from html import unescape
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

HTML = os.path.join(os.path.dirname(__file__), 'test-cases.html')
XLSX = os.path.join(os.path.dirname(__file__), 'test-cases.xlsx')

with open(HTML, 'r', encoding='utf-8') as f:
    src = f.read()

def strip_tags(s):
    s = re.sub(r'<br\s*/?>', '\n', s)
    s = re.sub(r'</li>\s*<li>', '\n• ', s)
    s = re.sub(r'<li>', '• ', s)
    s = re.sub(r'</li>', '', s)
    s = re.sub(r'</?(ol|ul|p|div|span|b|i|em|strong|code)[^>]*>', '', s)
    s = re.sub(r'<[^>]+>', '', s)
    s = unescape(s)
    s = re.sub(r'\n\s*\n', '\n', s)
    s = re.sub(r'[ \t]+', ' ', s)
    return s.strip()

# desc uses (.*?) not [^<]* because some descriptions contain inline tags
# (e.g. §16 has <code>…</code>, §17 has <b>…</b>); [^<]* would fail to close
# those sections and silently merge their cases into the previous module.
sections = list(re.finditer(r'<section class="module" id="([^"]+)">\s*<h2>([^<]+)<span class="pid">[^<]+</span></h2>\s*<p class="desc">(.*?)</p>', src, re.DOTALL))
section_boundaries = []
for i, m in enumerate(sections):
    start = m.start()
    end = sections[i+1].start() if i+1 < len(sections) else len(src)
    section_boundaries.append((start, end, strip_tags(m.group(2)), m.group(1)))

tc_pattern = re.compile(
    r'<div class="tc[^"]*"\s+data-type="([^"]+)"\s+data-pol="([^"]+)"\s+data-pri="([^"]+)"\s+data-role="([^"]+)">'
    r'\s*<div class="tc-head">.*?<span class="tc-id">([^<]+)</span>.*?</div>'
    r'\s*<div class="tc-title">([^<]+)</div>'
    r'\s*<div class="tc-body">(.*?)</div>'
    r'\s*</div>',
    re.DOTALL
)
generic_h3_pattern = re.compile(r'<h3[^>]*>(.*?)</h3>', re.DOTALL)

def find_subsection(pos):
    best = ''
    for m in generic_h3_pattern.finditer(src, 0, pos):
        best = strip_tags(m.group(1))
    return best

def find_module(pos):
    for start, end, title, _id in section_boundaries:
        if start <= pos < end:
            return title
    return ''

def parse_body(body):
    pre, steps, exp, brd = '', '', '', ''
    pre_m = re.search(r'<span class="label">Preconditions</span>\s*<ul>(.*?)</ul>', body, re.DOTALL)
    if pre_m: pre = strip_tags(pre_m.group(1))
    steps_m = re.search(r'<span class="label">Steps</span>\s*<ol>(.*?)</ol>', body, re.DOTALL)
    if steps_m: steps = strip_tags(steps_m.group(1))
    exp_m = re.search(r'<div class="exp">.*?<span class="label">Expected</span>\s*(.*?)</div>', body, re.DOTALL)
    if exp_m: exp = strip_tags(exp_m.group(1))
    brd_m = re.search(r'<p class="brd">(.*?)</p>', body, re.DOTALL)
    if brd_m: brd = strip_tags(brd_m.group(1))
    return pre, steps, exp, brd

type_label = {'func': 'Functional', 'ui': 'UI / UX', 'val': 'Validation', 'api': 'API',
              'db': 'Database', 'nf': 'Non-Functional', 'sec': 'Security'}
pol_label = {'pos': 'Positive', 'neg': 'Negative', 'bdy': 'Boundary'}
pri_label = {'p0': 'P0 — Blocker', 'p1': 'P1 — High', 'p2': 'P2 — Medium'}

# AUTO = deterministic, machine-assertable. SEMI = visual / experiential judgement.
AUTO_TYPES = {'func', 'val', 'api', 'db', 'sec'}
def exec_mode(code): return 'AUTO' if code in AUTO_TYPES else 'SEMI'

rows = []
for m in tc_pattern.finditer(src):
    pre, steps, exp, brd = parse_body(m.group(7))
    rows.append({
        'ID': m.group(5).strip(),
        'Module': find_module(m.start()),
        'Subsection': find_subsection(m.start()),
        'Title': m.group(6).strip(),
        'Type': type_label.get(m.group(1), m.group(1)),
        'Exec Mode': exec_mode(m.group(1)),
        'Polarity': pol_label.get(m.group(2), m.group(2)),
        'Priority': pri_label.get(m.group(3), m.group(3)),
        'Role': m.group(4).strip(),
        'Preconditions': pre,
        'Steps': steps,
        'Expected Result': exp,
        'BRD Ref': brd,
        'Status': '',
        'Actual Result': '',
        'Severity': '',
        'Comments': '',
    })

print(f'Parsed {len(rows)} test cases')

headers = ['ID','Module','Subsection','Title','Type','Exec Mode','Polarity','Priority','Role',
           'Preconditions','Steps','Expected Result','BRD Ref',
           'Status','Actual Result','Severity','Comments']

# ---- styling helpers --------------------------------------------------------
header_fill = PatternFill('solid', fgColor='0E4D4E')
header_font = Font(name='Georgia', bold=True, color='FBF5E5', size=11)
body_font = Font(name='Georgia', size=10)
thin = Side(border_style='thin', color='E8DFC9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical='top', wrap_text=True)
center = Alignment(vertical='top', horizontal='center', wrap_text=True)
mode_fill = {'AUTO': PatternFill('solid', fgColor='D9E8E5'), 'SEMI': PatternFill('solid', fgColor='F5E2D6')}

CENTER_COLS = {5, 6, 7, 8, 14, 16}  # Type, Exec Mode, Polarity, Priority, Status, Severity

wb = Workbook()
ws = wb.active
ws.title = 'Test Cases'
ws.append(headers)
for col in range(1, len(headers)+1):
    c = ws.cell(row=1, column=col)
    c.fill = header_fill; c.font = header_font; c.border = border
    c.alignment = Alignment(vertical='center', horizontal='left')

for r in rows:
    ws.append([r[h] for h in headers])
    rnum = ws.max_row
    for col in range(1, len(headers)+1):
        c = ws.cell(row=rnum, column=col)
        c.font = body_font; c.border = border
        c.alignment = center if col in CENTER_COLS else wrap
    ws.cell(row=rnum, column=6).fill = mode_fill.get(r['Exec Mode'])  # colour Exec Mode

# Dropdowns for the QA to fill during execution
status_dv = DataValidation(type='list', formula1='"Pass,Fail,Blocked,Not Run,N/A"', allow_blank=True)
sev_dv = DataValidation(type='list', formula1='"Critical,High,Medium,Low"', allow_blank=True)
ws.add_data_validation(status_dv); ws.add_data_validation(sev_dv)
last = ws.max_row
status_dv.add(f'N2:N{last}')   # Status
sev_dv.add(f'P2:P{last}')      # Severity

widths = {'A':24,'B':26,'C':28,'D':56,'E':14,'F':11,'G':12,'H':14,'I':16,
          'J':40,'K':58,'L':56,'M':22,'N':12,'O':40,'P':12,'Q':30}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 28
ws.freeze_panes = 'D2'
ws.auto_filter.ref = ws.dimensions

# ---- Bugs Only sheet (populated by _merge_results.py after a run) -----------
bugs = wb.create_sheet('Bugs Only')
bug_headers = ['Bug ID','Test Case ID','Module','Title','Severity','Priority',
               'Steps to Reproduce','Expected','Actual','Screenshot','BRD Ref',
               'Status','Assigned To','Notes']
bugs.append(['Bugs Only — auto-populated from failed/blocked cases after execution. Empty until a run is merged.'])
bugs['A1'].font = Font(name='Georgia', italic=True, size=10, color='8A7A55')
bugs.append(bug_headers)
for col in range(1, len(bug_headers)+1):
    c = bugs.cell(row=2, column=col)
    c.fill = header_fill; c.font = header_font; c.border = border
    c.alignment = Alignment(vertical='center', horizontal='left')
bug_widths = {'A':12,'B':22,'C':24,'D':54,'E':12,'F':14,'G':56,'H':54,'I':44,'J':20,'K':22,'L':12,'M':16,'N':30}
for col, w in bug_widths.items():
    bugs.column_dimensions[col].width = w
bugs.freeze_panes = 'A3'

# ---- Summary ----------------------------------------------------------------
summary = wb.create_sheet('Summary')
summary.append(['PRIME Rural — UAT Execution Coverage'])
summary['A1'].font = Font(name='Georgia', bold=True, size=16, color='0E4D4E')

from collections import Counter
mod, typ, mode, pol, pri, role = (Counter() for _ in range(6))
for r in rows:
    mod[r['Module']] += 1; typ[r['Type']] += 1; mode[r['Exec Mode']] += 1
    pol[r['Polarity']] += 1; pri[r['Priority']] += 1; role[r['Role']] += 1

def write_block(start_row, title, d):
    summary.cell(row=start_row, column=1, value=title).font = Font(name='Georgia', bold=True, size=12, color='0E4D4E')
    rr = start_row + 1
    summary.cell(row=rr, column=1, value='Bucket').font = Font(name='Georgia', bold=True)
    summary.cell(row=rr, column=2, value='Count').font = Font(name='Georgia', bold=True)
    rr += 1
    for k, v in sorted(d.items(), key=lambda kv: -kv[1]):
        summary.cell(row=rr, column=1, value=k or '(blank)').font = Font(name='Georgia')
        summary.cell(row=rr, column=2, value=v).font = Font(name='Georgia')
        rr += 1
    return rr + 1

summary.cell(row=2, column=1, value=f'Total test cases: {len(rows)}').font = Font(name='Georgia', bold=True, size=14)
row = 4
row = write_block(row, 'By Execution Mode', mode)
row = write_block(row, 'By Type', typ)
row = write_block(row, 'By Priority', pri)
row = write_block(row, 'By Module', mod)
row = write_block(row, 'By Polarity', pol)
row = write_block(row, 'By Role', role)
summary.column_dimensions['A'].width = 50
summary.column_dimensions['B'].width = 12

wb.move_sheet('Summary', offset=-(wb.sheetnames.index('Summary')))  # Summary first
wb.save(XLSX)
print(f'Wrote {XLSX}  (AUTO={mode["AUTO"]}, SEMI={mode["SEMI"]})')
