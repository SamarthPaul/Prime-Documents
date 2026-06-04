# -*- coding: utf-8 -*-
"""Append Journey-1 (User Management / RBAC) bugs to the Internal QA tracker.
Found 4 Jun 2026 via live MCP scout of staging. Appends after the last filled
row, copies the colour-less style of the row above (borders/wrap), no merges,
no colour, no invented columns. Idempotent-ish: re-running appends duplicates,
so run once.
"""
import copy
import datetime
import openpyxl

FILE = "Prime_Rural_Bug_Tracker_updated.xlsx"
RAISED_BY = "Samarth Paul"
RAISE_DATE = datetime.datetime(2026, 6, 4)

# (Module C, Type D, Testing E, UserId F, Location G, Interface H, Short I, Long J, Intended K, Priority M)
BUGS = [
    ("User Management → SPA Session", "Bug", "Functional Testing", "Admin / Core Team IT / Field",
     "Sidebar / User Menu (all pages)", "Web",
     "Stale user boot after re-login — previous user's name + nav shown until manual reload",
     "After signing out and logging in as a different user in the same browser session, the SPA keeps "
     "showing the PREVIOUS user's display name AND navigation menu until a full page reload. Reproduced "
     "3x: logging in as a Field user after a Core Team IT session showed \"CoreTeamTest\" + Core-Team nav; "
     "logging in as Core Team IT after a Field session showed \"Samarth Field User\" + the field-restricted "
     "nav. A hard reload corrects both name and nav. Server-side roles/permissions are correct; only the "
     "client boot is stale. Risk: identity confusion + momentary display of nav a role should not see, "
     "especially on shared field devices.",
     "On login the SPA must refetch the user boot so the newly logged-in user's name and role-appropriate "
     "navigation render immediately, with no manual reload.",
     "P2"),

    ("User Management → Role Profiles", "Bug", "Functional Testing", "Core Team IT",
     "User Management → Create User → Role Profile Assignment", "Web",
     "Duplicate Core Team IT role profile — 'Core Team Information Technology' exists alongside 'Core Team - IT'",
     "The Role Profile Assignment on the create-user form lists both \"Core Team - IT\" and \"Core Team "
     "Information Technology\". Per the BRD Role & Permission Matrix there is a single Core Team IT role. The "
     "duplicate is confusing and risks inconsistent permissions depending on which card is picked.",
     "Only one canonical \"Core Team - IT\" role profile should exist; remove \"Core Team Information "
     "Technology\". Role-profile names must match the BRD.",
     "P2"),

    ("User Management → Role Profiles", "Bug", "Functional Testing", "Core Team IT",
     "User Management → Create User → Role Profile Assignment", "Web",
     "'Viewer' role profile should be 'District Coordinator' (district-scoped); no generic Viewer in BRD",
     "The create-user form offers a \"Viewer\" role profile that does not exist in the BRD. Per the BRD there "
     "is no generic Viewer role; there is a \"District Coordinator\" who has view access to all workspaces but "
     "scoped to the coordinator's assigned district (the district is set while creating the user).",
     "Replace \"Viewer\" with \"District Coordinator\": view-only across all workspaces, scoped to the "
     "coordinator's assigned district, with the district assigned at user creation.",
     "P2"),
]

wb = openpyxl.load_workbook(FILE)
ws = wb["Internal QA"]

# find last filled data row (col C = Module)
last = 0
for r in range(3, ws.max_row + 1):
    if ws.cell(row=r, column=3).value:
        last = r
start = last + 1
TEMPLATE = last  # copy style from the last filled row (colour-less)
COLS = list(range(1, 23))  # A..V

for i, b in enumerate(BUGS):
    r = start + i
    module, typ, testing, uid, loc, iface, short, long, intended, prio = b
    for c in COLS:
        ws.cell(row=r, column=c)._style = copy.copy(ws.cell(row=TEMPLATE, column=c)._style)
    ws.cell(row=r, column=2, value="Staging")
    ws.cell(row=r, column=3, value=module)
    ws.cell(row=r, column=4, value=typ)
    ws.cell(row=r, column=5, value=testing)
    ws.cell(row=r, column=6, value=uid)
    ws.cell(row=r, column=7, value=loc)
    ws.cell(row=r, column=8, value=iface)
    ws.cell(row=r, column=9, value=short)
    ws.cell(row=r, column=10, value=long)
    ws.cell(row=r, column=11, value=intended)
    ws.cell(row=r, column=13, value=prio)
    ws.cell(row=r, column=15, value=RAISED_BY)
    ws.cell(row=r, column=16, value=RAISE_DATE)
    ws.cell(row=r, column=17, value="Open")
    ws.cell(row=r, column=19, value="Open")
    ws.cell(row=r, column=21, value="Open")

wb.save(FILE)
print(f"Appended {len(BUGS)} bugs at rows {start}-{start + len(BUGS) - 1}")
