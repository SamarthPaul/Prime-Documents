"""
Build the requirement-traceability gap matrix (HTML + xlsx) from the
journey-doc-anchored audit (frameworks + EP Profile + Product form).

Source of truth = journey doc + wireframe HTML + BRD + repo (see memory
source-of-truth-frameworks). Each ROW records, per rule:
  area, module, type, detail, brd (Y/P/N), tests (Y/P/N), flag, note
where flag ∈ '' | 'stale' (test contradicts source) | 'gap' (real coverage
gap) | 'brd' (BRD defect vs repo/journey) | 'build' (not built in repo).

Rebuild:  python _build_traceability.py
Outputs:  traceability-matrix.html  +  traceability-matrix.xlsx
"""
import os, html
from collections import OrderedDict
from _traceability_data import ROWS

HERE = os.path.dirname(__file__)
COLS = ['area', 'module', 'type', 'detail', 'brd', 'tests', 'flag', 'note']
def R(t): return dict(zip(COLS, t))
rows = [R(t) for t in ROWS]

def cov(v): return 1 if v in ('Y', 'P') else 0

# ---- group by area -> module ------------------------------------------------
areas = OrderedDict()
for r in rows:
    areas.setdefault(r['area'], OrderedDict()).setdefault(r['module'], []).append(r)

# ---- HTML -------------------------------------------------------------------
def esc(s): return html.escape(str(s or ''))
BADGE = {'Y': ('✓', 'y'), 'P': ('~', 'p'), 'N': ('✗', 'n')}
FLAGTXT = {'stale': 'STALE TEST', 'gap': 'NOT TESTED', 'brd': 'BRD DEFECT', 'build': 'NOT BUILT'}

tot = len(rows)
brd_ok = sum(cov(r['brd']) for r in rows)
tst_ok = sum(cov(r['tests']) for r in rows)
n_stale = sum(1 for r in rows if r['flag'] == 'stale')
n_gap = sum(1 for r in rows if r['flag'] == 'gap' or (r['tests'] == 'N' and not r['flag']))
n_brd = sum(1 for r in rows if r['flag'] in ('brd', 'build'))

p = ['''<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>PRIME Rural — Requirement Traceability Matrix</title>
<style>
:root{--teal:#0E4D4E;--sea:#1A6B6D;--terracotta:#C26849;--cream:#FBF5E5;--offwhite:#FDFAF0;--ink:#2A2520;--ink-soft:#6B6359;--border:#E8DFC9}
*{box-sizing:border-box}body{font-family:Georgia,'Times New Roman',serif;color:var(--ink);background:var(--cream);margin:0;line-height:1.5}
.wrap{max-width:1200px;margin:0 auto;padding:28px 22px 80px}
h1{color:var(--teal);font-size:30px;margin:0 0 4px}.sub{color:var(--ink-soft);margin:0 0 16px}
.note{background:var(--offwhite);border-left:4px solid var(--terracotta);padding:12px 16px;border-radius:6px;margin:14px 0}
h2{color:var(--teal);font-size:22px;border-left:4px solid var(--terracotta);padding-left:12px;margin:34px 0 6px}
h3{color:var(--sea);font-size:16px;margin:18px 0 4px}
table{width:100%;border-collapse:collapse;background:#fff;border:1px solid var(--border);border-radius:8px;overflow:hidden;margin-bottom:8px}
th{background:var(--teal);color:var(--cream);text-align:left;padding:7px 9px;font-size:12px}
td{border-top:1px solid var(--border);padding:7px 9px;font-size:12.5px;vertical-align:top}
.ty{color:var(--sea);font-size:11px;text-transform:uppercase;letter-spacing:.4px;white-space:nowrap}
.b{text-align:center;font-weight:bold;width:34px}
.y{color:#2e7d32}.p{color:var(--terracotta)}.n{color:#b3402a}
.tally{color:var(--ink-soft);font-size:12.5px;margin:2px 0 6px}
.flag{font-size:10px;font-weight:bold;padding:1px 6px;border-radius:8px;white-space:nowrap}
.f-stale{background:#f3d9cf;color:#9a3412}.f-gap{background:#efe7cf;color:#7a6320}.f-brd{background:#e7d6ef;color:#6b2f8a}.f-build{background:#d9e2ef;color:#274690}
tr.stale td{background:#fdf3ee}tr.brd td,tr.build td{background:#faf4fb}
.sumtab td,.sumtab th{font-size:12.5px}
.bar{display:inline-block;height:9px;border-radius:5px;background:var(--sea);vertical-align:middle}
.barbg{display:inline-block;width:90px;height:9px;border-radius:5px;background:#e8dfc9;vertical-align:middle;margin-right:6px}
.toolbar{position:sticky;top:0;background:var(--cream);padding:10px 0;z-index:5;border-bottom:1px solid var(--border)}
.btn{font-family:inherit;font-size:12px;border:1px solid var(--sea);background:#fff;color:var(--sea);padding:5px 12px;border-radius:16px;cursor:pointer;margin-right:6px}
.btn.on{background:var(--sea);color:#fff}
@media print{.toolbar{display:none}th{background:#0E4D4E!important;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
</style></head><body><div class="wrap">
<h1>PRIME Rural — Requirement Traceability Matrix</h1>
<p class="sub">Source of truth: <b>user journey doc</b> + <b>wireframe HTML</b> + <b>comprehensive BRD</b> (+ product repo for field/validation logic). Frameworks · Entrepreneur Profile · Product form.</p>
<div class="note"><b>How to read:</b> each row is one rule (dropdown / validation / skip-logic / calc / journey / auto-fill) extracted from the source-of-truth. <b>In BRD</b> and <b>In Tests</b> show whether it is documented / covered (✓ full · ~ partial · ✗ missing).
<span class="flag f-stale">STALE TEST</span> = a test case that contradicts the source (rewrite it); <span class="flag f-gap">NOT TESTED</span> = real coverage gap; <span class="flag f-brd">BRD DEFECT</span> / <span class="flag f-build">NOT BUILT</span> = BRD contradicts repo / feature absent in repo.</div>
''']
p.append(f'''<div class="note" style="border-color:var(--sea)"><b>Headline:</b> {tot} rules audited &nbsp;·&nbsp; in BRD <b>{brd_ok}</b> ({100*brd_ok//tot}%) &nbsp;·&nbsp; curated-test coverage <b>{tst_ok}</b> ({100*tst_ok//tot}%) &nbsp;·&nbsp; <b style="color:#9a3412">{n_stale}</b> stale tests (now realigned) &nbsp;·&nbsp; <b style="color:#7a6320">{n_gap}</b> untested rules &nbsp;·&nbsp; <b style="color:#6b2f8a">{n_brd}</b> BRD/build defects.</div>''')
p.append('''<div class="note"><b>Remediation status:</b> the stale tests have been rewritten in <b>test-cases.html</b> to match the source-of-truth; a generated <b>rule-coverage suite</b> (PR-TC-CVR-*) now adds one executable check per uncovered rule, so the test workbook (<b>test-cases.xlsx</b>) covers every rule pending execution. Remaining <b>BRD/build defects</b> below are the reconciliation worklist for BA/dev (repo-vs-spec decisions are flagged, not unilaterally rewritten).</div>''')

p.append('<div class="toolbar"><button class="btn on" data-f="all" onclick="flt(this)">All rules</button>'
         '<button class="btn" data-f="gap" onclick="flt(this)">Gaps only</button>'
         '<button class="btn" data-f="stale" onclick="flt(this)">Stale tests</button>'
         '<button class="btn" data-f="brd" onclick="flt(this)">BRD/build defects</button></div>')

# ---- per-area summary table -------------------------------------------------
for area, mods in areas.items():
    p.append(f'<h2>{esc(area)}</h2>')
    # summary
    p.append('<table class="sumtab"><tr><th>Module</th><th>Rules</th><th>In BRD</th><th>In Tests</th><th>Test coverage</th><th>Flags</th></tr>')
    for mod, mr in mods.items():
        n = len(mr); b = sum(cov(x['brd']) for x in mr); t = sum(cov(x['tests']) for x in mr)
        st = sum(1 for x in mr if x['flag'] == 'stale'); de = sum(1 for x in mr if x['flag'] in ('brd', 'build'))
        pct = 100*t//n if n else 0
        fl = []
        if st: fl.append(f'<span class="flag f-stale">{st} stale</span>')
        if de: fl.append(f'<span class="flag f-brd">{de} defect</span>')
        p.append(f'<tr><td><b>{esc(mod)}</b></td><td>{n}</td><td>{b}</td><td>{t}</td>'
                 f'<td><span class="barbg"><span class="bar" style="width:{int(0.9*pct)}px"></span></span>{pct}%</td>'
                 f'<td>{" ".join(fl)}</td></tr>')
    p.append('</table>')
    # detail tables
    for mod, mr in mods.items():
        n = len(mr); b = sum(cov(x['brd']) for x in mr); t = sum(cov(x['tests']) for x in mr)
        p.append(f'<h3>{esc(mod)}</h3>')
        p.append(f'<p class="tally">Rules: <b>{n}</b> · In BRD: {b} · In Tests: {t}</p>')
        p.append('<table><tr><th>Type</th><th>Rule</th><th class="b">BRD</th><th class="b">Tests</th><th>Flag / gap note</th></tr>')
        for x in mr:
            cls = x['flag'] if x['flag'] in ('stale', 'brd', 'build') else ''
            bb = BADGE[x['brd']]; tb = BADGE[x['tests']]
            fl = ''
            if x['flag']:
                fc = {'stale': 'f-stale', 'gap': 'f-gap', 'brd': 'f-brd', 'build': 'f-build'}[x['flag']]
                fl = f'<span class="flag {fc}">{FLAGTXT[x["flag"]]}</span> '
            elif x['tests'] == 'N':
                fl = '<span class="flag f-gap">NOT TESTED</span> '
            p.append(f'<tr class="{cls}" data-flag="{x["flag"] or ("gap" if x["tests"]=="N" else "")}">'
                     f'<td class="ty">{esc(x["type"])}</td><td>{esc(x["detail"])}</td>'
                     f'<td class="b {bb[1]}">{bb[0]}</td><td class="b {tb[1]}">{tb[0]}</td>'
                     f'<td>{fl}{esc(x["note"])}</td></tr>')
        p.append('</table>')

p.append('''<script>
function flt(btn){document.querySelectorAll('.btn').forEach(b=>b.classList.remove('on'));btn.classList.add('on');
var f=btn.dataset.f;document.querySelectorAll('tbody tr,table tr').forEach(function(tr){
 if(!tr.dataset.flag&&tr.querySelector('th')) return;
 var fg=tr.dataset.flag||'';var show=(f==='all')||(f==='gap'&&(fg==='gap'))||(f==='stale'&&fg==='stale')||(f==='brd'&&(fg==='brd'||fg==='build'));
 if(tr.querySelector('td')&&tr.children.length>=5&&tr.querySelector('.ty')!==null){tr.style.display=show?'':'none';}});}
</script>''')
p.append('</div></body></html>')
open(os.path.join(HERE, 'traceability-matrix.html'), 'w', encoding='utf-8', newline='\n').write('\n'.join(p))

# ---- xlsx -------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
wb = Workbook()
teal = PatternFill('solid', fgColor='0E4D4E'); whitef = Font(color='FBF5E5', bold=True)
def sheet(name, headers, data):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for c in ws[1]: c.fill = teal; c.font = whitef
    for d in data: ws.append(d)
    ws.freeze_panes = 'A2'
    widths = {'Area': 18, 'Module': 22, 'Type': 12, 'Rule': 70, 'In BRD': 8, 'In Tests': 9, 'Flag': 12, 'Note': 60}
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = widths.get(h, 16)
    for r_ in ws.iter_rows(min_row=2):
        for c in r_: c.alignment = Alignment(vertical='top', wrap_text=True)
    return ws
wb.remove(wb.active)
FLABEL = {'': '', 'stale': 'STALE TEST', 'gap': 'NOT TESTED', 'brd': 'BRD DEFECT', 'build': 'NOT BUILT'}
matrix = [[r['area'], r['module'], r['type'], r['detail'], r['brd'], r['tests'],
           FLABEL[r['flag']] or ('NOT TESTED' if r['tests'] == 'N' else ''), r['note']] for r in rows]
sheet('Matrix', ['Area', 'Module', 'Type', 'Rule', 'In BRD', 'In Tests', 'Flag', 'Note'], matrix)
# summary
summ = []
for area, mods in areas.items():
    for mod, mr in mods.items():
        n = len(mr); b = sum(cov(x['brd']) for x in mr); t = sum(cov(x['tests']) for x in mr)
        st = sum(1 for x in mr if x['flag'] == 'stale'); de = sum(1 for x in mr if x['flag'] in ('brd', 'build'))
        summ.append([area, mod, n, b, t, f'{100*t//n if n else 0}%', st, de])
sheet('Summary', ['Area', 'Module', 'Rules', 'In BRD', 'In Tests', 'Test %', 'Stale', 'Defects'], summ)
sheet('Gaps Only', ['Area', 'Module', 'Type', 'Rule', 'In BRD', 'In Tests', 'Flag', 'Note'],
      [m for m, r in zip(matrix, rows) if r['tests'] == 'N' or r['flag'] == 'gap'])
sheet('Stale Tests', ['Area', 'Module', 'Type', 'Rule', 'In BRD', 'In Tests', 'Flag', 'Note'],
      [m for m, r in zip(matrix, rows) if r['flag'] == 'stale'])
sheet('BRD & Build Defects', ['Area', 'Module', 'Type', 'Rule', 'In BRD', 'In Tests', 'Flag', 'Note'],
      [m for m, r in zip(matrix, rows) if r['flag'] in ('brd', 'build')])
wb.save(os.path.join(HERE, 'traceability-matrix.xlsx'))

print(f'Wrote traceability-matrix.html + .xlsx')
print(f'  {tot} rules | BRD {brd_ok} ({100*brd_ok//tot}%) | Tests {tst_ok} ({100*tst_ok//tot}%) | stale {n_stale} | untested {n_gap} | BRD/build defects {n_brd}')
