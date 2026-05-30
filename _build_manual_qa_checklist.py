"""
Build generated/manual-qa-checklist.html from the Blocked cases in
test-cases.xlsx — the items that the automated + visual passes could NOT verify
and that need a human (or dev input-contracts). Grouped into tester-friendly
sections with a Pass/Fail/Notes column. Monsoon Court styled (matches test-cases.html).
"""
import os, html
from collections import OrderedDict
from openpyxl import load_workbook

HERE = os.path.dirname(__file__)
wb = load_workbook(os.path.join(HERE, 'test-cases.xlsx'))
tc = wb['Test Cases']
hdr = {c.value: i for i, c in enumerate(tc[1])}

def cell(row, name):
    return row[hdr[name]].value or ''

# collect items still needing verification: Blocked OR Not-Run (blank). Exclude Pass/Fail/Flaky.
rows = []
for r in tc.iter_rows(min_row=2):
    st = (r[hdr['Status']].value or '').strip()
    if st in ('Pass', 'Fail', 'Flaky'):
        continue
    d = {k: cell(r, k) for k in ['ID', 'Module', 'Subsection', 'Title', 'Type', 'Priority', 'Actual Result', 'Expected Result']}
    d['_state'] = 'Blocked (triaged)' if st == 'Blocked' else 'Not run'
    rows.append(d)

# ---- categorise by module group (reliable) ----------------------------------
def modnum(m):
    try: return int(str(m).split('.')[0])
    except: return 99

def categorise(x):
    n = modnum(x['Module'])
    g = {
        19: '1. Dashboards & Maps — visual / interaction (hover, drawer, drill-down)',
        15: '2. Entrepreneur Profile — UI flows (GPS, photo, KYC, selection, prioritisation, networking)',
        17: '3. Framework calc / UI deltas (per-framework — UI form-fill or DEV calc input-contracts)',
        16: '4. Product Profile UI (process-flow drag, SKU, variants, drawer)',
        14: '5. Masters & reference data (cascades, CSV import, archive UI)',
        18: '6. Activity Logger (calendar / board UI)',
        20: '7. Notifications (trigger events, bell, deep-links)',
        11: '8. Error handling & empty states (UI)',
        13: '9. User Management (CSV import, audit, filters UI)',
    }.get(n)
    if g: return g
    if n in (5, 6): return '10. Browser, device & responsive (cross-browser, zoom, heavy-data)'
    if n == 9: return '11. Accessibility (keyboard, focus-trap, screen-reader, charts)'
    if n in (8, 12): return '12. UI/UX & locale (visual / format checks)'
    if n in (4, 7): return '13. Entry / Auth (remaining edge cases)'
    if n == 10: return '14. Performance (memory, N+1 — profiling)  [DEV]'
    if n == 22: return '15. Database (indexes, EXPLAIN, encryption, backups)  [DEV/INFRA]'
    if n == 23: return '16. Security (VAPT, file-upload, audit log)  [DEV/EXTERNAL]'
    if n == 21: return '17. API (rate-limit, disabled-user key)  [DEV/INFRA]'
    return '18. Other'

groups = OrderedDict()
for x in rows:
    groups.setdefault(categorise(x), []).append(x)
groups = OrderedDict(sorted(groups.items(), key=lambda kv: int(kv[0].split('.')[0])))

PRI = {'P0 — Blocker': 'p0', 'P1 — High': 'p1', 'P2 — Medium': 'p2'}

# ---- emit HTML --------------------------------------------------------------
def esc(s): return html.escape(str(s or ''))

parts = ['''<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>PRIME Rural — Manual QA Checklist</title>
<style>
:root{--teal:#0E4D4E;--sea:#1A6B6D;--terracotta:#C26849;--cream:#FBF5E5;--offwhite:#FDFAF0;--ink:#2A2520;--ink-soft:#6B6359;--border:#E8DFC9}
*{box-sizing:border-box}body{font-family:Georgia,'Times New Roman',serif;color:var(--ink);background:var(--cream);margin:0;line-height:1.5}
.wrap{max-width:1100px;margin:0 auto;padding:28px 22px 80px}
h1{color:var(--teal);font-size:30px;margin:0 0 4px}.sub{color:var(--ink-soft);margin:0 0 18px}
.note{background:var(--offwhite);border-left:4px solid var(--terracotta);padding:12px 16px;border-radius:6px;margin:14px 0}
h2{color:var(--teal);font-size:21px;border-left:4px solid var(--terracotta);padding-left:12px;margin:30px 0 10px}
table{width:100%;border-collapse:collapse;background:#fff;border:1px solid var(--border);border-radius:8px;overflow:hidden;margin-bottom:6px}
th{background:var(--teal);color:var(--cream);text-align:left;padding:8px 10px;font-size:13px}
td{border-top:1px solid var(--border);padding:8px 10px;font-size:13px;vertical-align:top}
.id{font-family:ui-monospace,Menlo,monospace;font-size:12px;color:var(--sea);white-space:nowrap}
.pri{font-weight:bold;font-size:11px}.p0{color:#b3402a}.p1{color:var(--terracotta)}.p2{color:var(--ink-soft)}
.box{font-size:18px;color:var(--ink-soft)}.res{width:120px;color:var(--ink-soft)}.notes{width:160px;color:var(--border)}
.dev{background:#F5E2D6}
.count{color:var(--ink-soft);font-size:13px}
@media print{th{background:#0E4D4E!important;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
</style></head><body><div class="wrap">
<h1>PRIME Rural — Manual QA Checklist</h1>
<p class="sub">Items the automated Playwright suite + MCP visual pass could not verify. Companion to <b>test-cases.html</b> / the <b>Bugs Only</b> sheet (58 automated findings already filed).</p>
<div class="note"><b>How to use:</b> work top-down. ✅ tick the box, mark <b>Result</b> = Pass / Fail, add Notes. Log any Fail into the bug tracker with the case ID. Sections tagged <b>[DEV]</b>/<b>[INFRA]</b>/<b>[EXTERNAL]</b> are not pure manual-QA — route to the dev/infra owner. <b>Section 3 (framework calcs)</b> is fastest if the dev provides the <b>calc input-contracts</b> (which child-table fields feed each computed field) — then they can be API-tested in bulk instead of hand-filled.</div>
''']
parts.append(f'<p class="count">Total items to verify: <b>{len(rows)}</b> across {len(groups)} sections.</p>')

for title, items in groups.items():
    parts.append(f'<h2>{esc(title)} <span class="count">({len(items)})</span></h2>')
    dev = '[DEV' in title or '[INFRA' in title or '[EXTERNAL' in title
    parts.append(f'<table class="{ "dev" if dev else "" }"><tr><th>✓</th><th>ID</th><th>Pri</th><th>State</th><th>What to verify</th><th>How / expected</th><th class="res">Result</th><th class="notes">Notes</th></tr>')
    for x in sorted(items, key=lambda z: (PRI.get(z['Priority'], 'p3'), z['ID'])):
        pri = PRI.get(x['Priority'], '')
        prilbl = (x['Priority'] or '').split('—')[0].strip()
        howto = x['Actual Result'] or x['Expected Result']  # triage note, else the spec expectation
        parts.append(
            f'<tr><td class="box">☐</td><td class="id">{esc(x["ID"])}</td>'
            f'<td class="pri {pri}">{esc(prilbl)}</td>'
            f'<td style="font-size:11px;color:var(--ink-soft)">{esc(x["_state"])}</td>'
            f'<td>{esc(x["Title"])}</td>'
            f'<td>{esc(howto)[:320]}</td>'
            f'<td class="res">Pass / Fail</td><td class="notes">&nbsp;</td></tr>')
    parts.append('</table>')

parts.append('</div></body></html>')
out = os.path.join(HERE, 'manual-qa-checklist.html')
open(out, 'w', encoding='utf-8', newline='\n').write('\n'.join(parts))
print(f'Wrote {out} — {len(rows)} items, {len(groups)} sections')
for t, it in groups.items():
    print(f'  {len(it):3}  {t}')
