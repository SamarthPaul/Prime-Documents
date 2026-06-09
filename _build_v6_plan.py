"""
Builds PREF_CRM_Web_Project_Plan_v6.xlsx from the Dhwani Setu PM template.

- 5 sheets: Header, Project Plan, Gantt Chart (summary), How to Use, Version Control
- 14 cols on Project Plan: standard 13 Dhwani cols + UAT Batch (col 10) for phased delivery
- Milestone (WBS L1) rows colored by status. Sub-rows white. No merged cells. No column-wide fills.
- AutoFilter on Project Plan rows so user can filter by status / owner / UAT batch.

Owner mapping (per user):
  Navneet  = Developer (frontend + backend)
  Bhushan  = Tech Lead (architecture, role/permission infra, code review-led tasks)
  Samarth  = PM (planning, client coordination, signoffs)
  Chitranshi / Samarth = Testing
  Prakash  = DevOps
  Client   = MBMA / Apoorva AI / client SMEs where applicable

Dates: backdated from current actual completion. Today = 9-Jun-2026 (v6.1).
UAT-1 internal E2E executed 7-8 Jun 2026 against staging (272 cases run, ~60 bugs).
v6.1 status reconciliation: User Management verified Completed; Prototyping (Bug 54)
and Milestone Tracker (Bug 55) marked In Progress after P1 regressions found in UAT-1.
"""

from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\Dhwani\OneDrive - Dhwani Rural Information Systems Pvt. Ltd\Work\prime-rural\generated\PREF_CRM_Web_Project_Plan_v6.xlsx"

D = date  # alias

# -------- Column definitions --------
HEADERS = [
    "S.No", "WBS", "Activity", "Duration (Days)", "Resources",
    "Dependency", "Owner", "Success Criteria", "Current Status",
    "UAT Batch", "Start Date", "End Date", "Actual Start Date", "Actual End Date",
]
WIDTHS = [6, 8, 60, 10, 20, 12, 22, 55, 18, 11, 13, 13, 14, 14]

# -------- Row data --------
# Each row: (level, wbs, activity, duration, resources, dependency, owner,
#            success_criteria, status, uat_batch, start, end, act_start, act_end)
# level: 1 = milestone (colored), 2 = sub-activity
# dates: None or date(); duration: int or None

ROWS = [
    # WBS 1 -- System Foundation (Pre-UAT, Completed)
    (1, "1", "System Foundation", 14, "DevOps / Tech Lead", "", "Prakash",
     "Staging environment live; auth + role infra ready for app development",
     "Completed", "Pre-UAT", D(2025,9,1), D(2025,9,14), D(2025,9,1), D(2025,9,14)),
    (2, "1.1", "Frappe v16 staging environment setup", 3, "DevOps", "", "Prakash",
     "stgprime-rural.dhwaniris.in live with HTTPS", "Completed", "Pre-UAT",
     D(2025,9,1), D(2025,9,3), D(2025,9,1), D(2025,9,3)),
    (2, "1.2", "GitHub repo + CI/CD hooks (dhwani-ris/frappe_prime_rural)", 2, "Backend Dev", "1.1", "Navneet",
     "Repo with pre-commit hooks + branch protections", "Completed", "Pre-UAT",
     D(2025,9,4), D(2025,9,5), D(2025,9,4), D(2025,9,5)),
    (2, "1.3", "API token authentication setup + testing", 1, "Backend Dev", "1.1", "Navneet",
     "Admin API calls succeed with token header", "Completed", "Pre-UAT",
     D(2025,9,6), D(2025,9,6), D(2025,9,6), D(2025,9,6)),
    (2, "1.4", "Google SSO configuration (Gmail-based users)", 2, "DevOps", "1.1", "Prakash",
     "Gmail users land in Frappe via OAuth without password", "Not Started", "UAT-1",
     D(2026,5,26), D(2026,5,28), None, None),
    (2, "1.5", "Welcome email custom template (set-password link)", 1, "Backend Dev", "1.4", "Navneet",
     "New users receive role-aware welcome email with set-password link", "Not Started", "UAT-1",
     D(2026,5,27), D(2026,5,28), None, None),

    # WBS 2 -- User Management (UAT-1)
    (1, "2", "User Management", 10, "Developer / Tech Lead", "1", "Navneet",
     "Core Team IT can add users, assign roles + blocks; users can log in",
     "In Progress", "UAT-1", D(2025,10,10), D(2026,5,28), D(2025,10,10), None),
    (2, "2.1", "Add User form (6 fields: name, email, phone, role, district, block)", 4, "Frontend Dev", "1", "Navneet",
     "Form validates + creates Frappe User + assigns role; BRD PR-XC-001 (UAT-1 verified: add/list/edit/deactivate PASS)",
     "Completed", "UAT-1", D(2025,10,10), D(2026,6,7), D(2025,10,10), D(2026,6,7)),
    (2, "2.2", "Role assignment logic (6 roles -> sidebar menu)", 5, "Tech Lead", "2.1", "Bhushan",
     "Each role lands on its own workspace; sidebar items hidden per role (UAT-1 verified for Field User nav)",
     "Completed", "UAT-1", D(2025,10,12), D(2026,6,7), D(2025,10,12), D(2026,6,7)),
    (2, "2.3", "Login flow: email + password", 2, "Backend Dev", "1", "Navneet",
     "Standard Frappe login active for non-Gmail users", "Completed", "UAT-1",
     D(2025,10,15), D(2025,10,16), D(2025,10,15), D(2025,10,16)),
    (2, "2.4", "Login flow: Google SSO for Gmail users", 2, "DevOps", "1.4", "Prakash",
     "Gmail users log in via Google button", "Not Started", "UAT-1",
     D(2026,5,27), D(2026,5,28), None, None),
    (2, "2.5", "Block assignment for Fessociates -> auto-fill District/Block", 2, "Backend Dev", "2.1", "Navneet",
     "Fessociate sees only their block's EPs; admin can override",
     "Not Started", "UAT-1", D(2026,5,27), D(2026,5,28), None, None),
    (2, "2.6", "District Coordinator: view-all edit-none permission set", 2, "Tech Lead", "2.2", "Bhushan",
     "DC can browse all data but Save buttons disabled per role rule",
     "Not Started", "UAT-1", D(2026,5,27), D(2026,5,28), None, None),
    (2, "2.7", "Designer role: restricted sidebar + upload-only permissions", 2, "Tech Lead", "2.2", "Bhushan",
     "Designer sees only Logo/Label/Brochure draft upload screens",
     "Not Started", "UAT-1", D(2026,5,27), D(2026,5,28), None, None),
    (2, "2.8", "SELCO Executive role: restricted to Infrastructure EE tab", 2, "Tech Lead", "2.2", "Bhushan",
     "SELCO sees only EE pricing entry + solarisation approval flow",
     "Not Started", "UAT-1", D(2026,5,27), D(2026,5,28), None, None),

    # WBS 3 -- Masters (UAT-1, Completed)
    (1, "3", "Masters", 18, "Developer / Tech Lead", "1", "Navneet",
     "All seed masters loaded + linked across the doctype graph",
     "Completed", "UAT-1", D(2025,9,15), D(2025,10,5), D(2025,9,15), D(2025,10,5)),
    (2, "3.1", "District + Block + Village master (Meghalaya hierarchy)", 4, "Backend Dev", "1", "Navneet",
     "All Meghalaya districts/blocks/villages loaded + cascading lookup live",
     "Completed", "UAT-1", D(2025,9,15), D(2025,9,18), D(2025,9,15), D(2025,9,18)),
    (2, "3.2", "Sector + Business Basket + Product Type masters", 2, "Backend Dev", "1", "Navneet",
     "Sector picker cascades to Business Basket then Product Type",
     "Completed", "UAT-1", D(2025,9,19), D(2025,9,20), D(2025,9,19), D(2025,9,20)),
    (2, "3.3", "Cohort master", 1, "Backend Dev", "1", "Navneet",
     "Cohort picker available on EP Profile", "Completed", "UAT-1",
     D(2025,9,21), D(2025,9,21), D(2025,9,21), D(2025,9,21)),
    (2, "3.4", "Packaging master (12 types: Glass Jar ... Other)", 1, "Backend Dev", "1", "Navneet",
     "All 12 packaging types selectable in Packaging framework",
     "Completed", "UAT-1", D(2025,9,22), D(2025,9,22), D(2025,9,22), D(2025,9,22)),
    (2, "3.5", "Registration master (20 registration types)", 3, "Backend Dev", "1", "Navneet",
     "All 20 registration types selectable in Registration framework",
     "Completed", "UAT-1", D(2025,9,23), D(2025,9,25), D(2025,9,23), D(2025,9,25)),
    (2, "3.6", "Test Type master (14 test types)", 1, "Backend Dev", "1", "Navneet",
     "Test type picker live in Product Testing framework", "Completed", "UAT-1",
     D(2025,9,26), D(2025,9,26), D(2025,9,26), D(2025,9,26)),
    (2, "3.7", "Lab master (12 labs)", 1, "Backend Dev", "1", "Navneet",
     "Lab picker live in Product Testing framework", "Completed", "UAT-1",
     D(2025,9,27), D(2025,9,27), D(2025,9,27), D(2025,9,27)),
    (2, "3.8", "Standardisation Type master (13 types)", 1, "Backend Dev", "1", "Navneet",
     "All 13 SOP types selectable in Standardization framework", "Completed", "UAT-1",
     D(2025,9,28), D(2025,9,28), D(2025,9,28), D(2025,9,28)),
    (2, "3.9", "Scheme master (17 govt schemes)", 2, "Backend Dev", "1", "Navneet",
     "All 17 schemes selectable in Funding framework", "Completed", "UAT-1",
     D(2025,9,29), D(2025,9,30), D(2025,9,29), D(2025,9,30)),
    (2, "3.10", "Role setup: 6 roles with permission rules", 5, "Tech Lead", "1", "Bhushan",
     "Core IT, Core Prog, Fessociate, Designer, SELCO, DC roles created with default permissions",
     "Completed", "UAT-1", D(2025,10,1), D(2025,10,5), D(2025,10,1), D(2025,10,5)),

    # WBS 4 -- Entrepreneur Profile (UAT-1)
    (1, "4", "Entrepreneur Profile", 31, "Developer", "3", "Navneet",
     "Fessociates can create + maintain entrepreneur records end-to-end",
     "In Progress", "UAT-1", D(2025,10,6), D(2026,5,28), D(2025,10,6), None),
    (2, "4.1", "EP doctype backend (all fields + naming rule)", 5, "Backend Dev", "3", "Navneet",
     "EP-{name}-##### naming; all fields validated server-side",
     "Completed", "UAT-1", D(2025,10,6), D(2025,10,10), D(2025,10,6), D(2025,10,10)),
    (2, "4.2", "Vue: Profile Dashboard (9 sub-components)", 9, "Frontend Dev", "4.1", "Navneet",
     "Single-page entrepreneur view with stats + product shelf + DF rollup",
     "Completed", "UAT-1", D(2025,10,12), D(2025,10,20), D(2025,10,12), D(2025,10,20)),
    (2, "4.3", "Vue: Selection tab (10 rating fields + Rural E-Champ)", 3, "Frontend Dev", "4.2", "Navneet",
     "Average rating auto-calculates; Rural E-Champ flag works",
     "Completed", "UAT-1", D(2025,10,21), D(2025,10,23), D(2025,10,21), D(2025,10,23)),
    (2, "4.4", "Vue: Prioritisation tab", 1, "Frontend Dev", "4.2", "Navneet",
     "Cards reorder on dashboard by priority value", "Completed", "UAT-1",
     D(2025,10,24), D(2025,10,24), D(2025,10,24), D(2025,10,24)),
    (2, "4.5", "Vue: Location & Contact + Geolocation pin", 3, "Frontend Dev", "4.2", "Navneet",
     "Map pin saved; reverse-geocode populates district/block",
     "Completed", "UAT-1", D(2025,10,25), D(2025,10,27), D(2025,10,25), D(2025,10,27)),
    (2, "4.6", "Vue: Documents tab + photo gallery", 1, "Frontend Dev", "4.2", "Navneet",
     "Documents upload + preview + lightbox gallery", "Completed", "UAT-1",
     D(2025,10,28), D(2025,10,28), D(2025,10,28), D(2025,10,28)),
    (2, "4.7", "Vue: Networking & Mentoring tab (inline intro + meeting log)", 7, "Frontend Dev", "4.2", "Navneet",
     "PR feat/networking-mentoring-v2 merged; auto-sync meeting counts to parent",
     "Completed", "UAT-1", D(2025,10,29), D(2025,11,5), D(2025,10,29), D(2025,11,5)),
    (2, "4.8", "Vue: Intervention Logs (read-only sync from Activity Logger)", 14, "Frontend Dev", "9", "Navneet",
     "Every Activity Logger entry surfaces on EP timeline read-only",
     "In Progress", "UAT-1", D(2025,11,6), D(2026,5,28), D(2025,11,6), None),

    # WBS 5 -- Product Profile (UAT-1)
    (1, "5", "Product Profile", 28, "Developer", "4", "Navneet",
     "All entrepreneur products captured with variants, process flow + production capacity",
     "In Progress", "UAT-1", D(2025,11,6), D(2026,5,28), D(2025,11,6), None),
    (2, "5.1", "Product doctype backend + child tables", 3, "Backend Dev", "4", "Navneet",
     "Product + variant + process step doctypes deployed",
     "Completed", "UAT-1", D(2025,11,6), D(2025,11,8), D(2025,11,6), D(2025,11,8)),
    (2, "5.2", "Vue: Product Details tab (12 fields)", 6, "Frontend Dev", "5.1", "Navneet",
     "Form populates + saves; product type integration live",
     "Completed", "UAT-1", D(2025,11,9), D(2025,11,15), D(2025,11,9), D(2025,11,15)),
    (2, "5.3", "Vue: Product Variants tab (SKU auto-gen + volume calc)", 7, "Frontend Dev", "5.1", "Navneet",
     "SKU = VARIANT-SUBVARIANT-WEIGHT-UNIT; volume = LxBxH auto",
     "Completed", "UAT-1", D(2025,11,16), D(2025,11,22), D(2025,11,16), D(2025,11,22)),
    (2, "5.4", "Vue: Process Flow tab (drag-drop reorderable steps)", 6, "Frontend Dev", "5.1", "Navneet",
     "Drag-drop reorders + saves; machine + flow master linked",
     "Completed", "UAT-1", D(2025,11,23), D(2025,11,28), D(2025,11,23), D(2025,11,28)),
    (2, "5.5", "Vue: Production Capacity tab (agri-conditional + monthly tracker)", 60, "Frontend Dev", "5.3", "Navneet",
     "Available Months x Variants matrix renders + saves per-variant qty",
     "In Progress", "UAT-1", D(2025,11,29), D(2026,5,28), D(2025,11,29), None),
    (2, "5.6", "Available Months checkbox logic -> tracker generation", 30, "Frontend Dev", "5.5", "Navneet",
     "Selecting months auto-creates corresponding columns in tracker",
     "In Progress", "UAT-1", D(2025,12,1), D(2026,5,28), D(2025,12,1), None),

    # WBS 6 -- Diagnostic Frameworks (27)
    (1, "6", "Diagnostic Frameworks (27)", 130, "Developer", "5", "Navneet",
     "All 27 framework forms deployed; entrepreneurs can be diagnosed end-to-end",
     "In Progress", "UAT-1", D(2025,12,1), D(2026,5,28), D(2025,12,1), None),
    (2, "6.1",  "Manufacturing - Raw Materials",      5, "Frontend Dev", "5", "Navneet", "Existing + Required tabs with Cost/Batch calc + Summary", "Completed", "UAT-1", D(2025,12,1),  D(2025,12,5),  D(2025,12,1),  D(2025,12,5)),
    (2, "6.2",  "Manufacturing - Machinery",          5, "Frontend Dev", "5", "Navneet", "22 fields + 7 calcs + depreciation + solarisation flags", "Completed", "UAT-1", D(2025,12,6),  D(2025,12,10), D(2025,12,6),  D(2025,12,10)),
    (2, "6.3",  "Manufacturing - Human Resource",     4, "Frontend Dev", "5", "Navneet", "3 wage formulas + GoM compliance warnings + 8 intervention types", "Completed", "UAT-1", D(2025,12,11), D(2025,12,14), D(2025,12,11), D(2025,12,14)),
    (2, "6.4",  "Pricing - Unit Pricing",             8, "Frontend Dev", "6.1,6.2,6.3", "Navneet", "9-section cost buildup + 3 pricing methods x 2 channels", "Completed", "UAT-1", D(2025,12,15), D(2025,12,22), D(2025,12,15), D(2025,12,22)),
    (2, "6.5",  "Market - Market Research",           158, "Frontend Dev", "5", "Navneet", "4 sections + competitive pricing + SWOT (auto-avg cross-feed pending)", "In Progress", "UAT-1", D(2025,12,23), D(2026,5,28), D(2025,12,23), None),
    (2, "6.6",  "Market - Customer Analysis",         4, "Frontend Dev", "5", "Navneet", "6 segmentations + Existing/Required intervention toggles", "Completed", "UAT-1", D(2025,12,26), D(2025,12,29), D(2025,12,26), D(2025,12,29)),
    (2, "6.7",  "Market - Market Linkage",            4, "Frontend Dev", "5", "Navneet", "Channel revenue tracking + annualized revenue calc", "Completed", "UAT-1", D(2025,12,30), D(2026,1,2),   D(2025,12,30), D(2026,1,2)),
    (2, "6.8",  "Market - Promotions",                3, "Frontend Dev", "5", "Navneet", "Existing + Required table with channel/mode/platform + frequency", "Completed", "UAT-1", D(2026,1,3),   D(2026,1,5),   D(2026,1,3),   D(2026,1,5)),
    (2, "6.9",  "Product Dev - Registration & Licensing", 5, "Frontend Dev", "5", "Navneet", "20-type master + Lifetime Validity conditional + status chips + 5-status pipeline", "Completed", "UAT-1", D(2026,1,6),   D(2026,1,10),  D(2026,1,6),   D(2026,1,10)),
    (2, "6.10", "Product Dev - Prototyping",          4, "Frontend Dev", "5", "Navneet", "Conditional dropdown + iteration child table [Bug 54 P1: cannot save - server 500 on framework_name; REGRESSED in UAT-1]", "In Progress", "UAT-1", D(2026,1,11),  D(2026,1,14),  D(2026,1,11),  None),
    (2, "6.11", "Product Dev - Testing",              4, "Frontend Dev", "5", "Navneet", "14 test types + 12 labs + consent management + date chain validation", "Completed", "UAT-1", D(2026,1,15),  D(2026,1,18),  D(2026,1,15),  D(2026,1,18)),
    (2, "6.12", "Product Dev - Standardization",      4, "Frontend Dev", "5", "Navneet", "13 SOP types + auto Product Sector + version tracking", "Completed", "UAT-1", D(2026,1,19),  D(2026,1,22),  D(2026,1,19),  D(2026,1,22)),
    (2, "6.13", "Product Dev - Quality Control",      6, "Frontend Dev", "5", "Navneet", "23-question star rating + average + maturity tier + distribution chart", "Completed", "UAT-1", D(2026,1,23),  D(2026,1,28),  D(2026,1,23),  D(2026,1,28)),
    (2, "6.14", "Product Dev - Packaging",            7, "Frontend Dev", "5", "Navneet", "Per-variant detail + monthly qty matrix + sealing multi-select", "Completed", "UAT-1", D(2026,1,29),  D(2026,2,4),   D(2026,1,29),  D(2026,2,4)),
    (2, "6.15", "Marketing Tools - Brand Name",       4, "Frontend Dev", "5", "Navneet", "Dual approval flow + market scope chips + log book", "Completed", "UAT-1", D(2026,2,5),   D(2026,2,8),   D(2026,2,5),   D(2026,2,8)),
    (2, "6.16", "Marketing Tools - Logo",             3, "Frontend Dev", "6.15", "Navneet", "Designer assignment + 3 drafts + dual approval", "Completed", "UAT-1", D(2026,2,9),   D(2026,2,11),  D(2026,2,9),   D(2026,2,11)),
    (2, "6.17", "Marketing Tools - Business Card",    3, "Frontend Dev", "6.15", "Navneet", "Auto-populated content + QR uploads", "Completed", "UAT-1", D(2026,2,12),  D(2026,2,14),  D(2026,2,12),  D(2026,2,14)),
    (2, "6.18", "Marketing Tools - Label",            102, "Frontend Dev", "5", "Navneet", "Per-variant table + AI Generate storyline (Apoorva dep) + 5-status pipeline", "In Progress", "UAT-1", D(2026,2,15),  D(2026,5,28),  D(2026,2,15), None),
    (2, "6.19", "Marketing Tools - Brochure",         5, "Frontend Dev", "5", "Navneet", "Brochure setup + product listings + status + log book", "Completed", "UAT-1", D(2026,2,22),  D(2026,2,26),  D(2026,2,22),  D(2026,2,26)),
    (2, "6.20", "Skilling - Digital Literacy",        4, "Frontend Dev", "5", "Navneet", "Access & Devices + 9-platform star rating + intervention flags", "Completed", "UAT-1", D(2026,2,27),  D(2026,3,2),   D(2026,2,27),  D(2026,3,2)),
    (2, "6.21", "Skilling - Capacity Building",       6, "Frontend Dev", "5", "Navneet", "Scorecard Gap calc + Past Trainings (6-dim stars) + Required gap analysis", "Completed", "UAT-1", D(2026,3,3),   D(2026,3,8),   D(2026,3,3),   D(2026,3,8)),
    (2, "6.22", "Skilling - Business Plan Canvas",    3, "Frontend Dev", "5", "Navneet", "9 Canvas sections + SWOT (BPC HTML pending from client)", "Completed", "UAT-1", D(2026,3,9),   D(2026,3,11),  D(2026,3,9),   D(2026,3,11)),
    (2, "6.23", "Logistics - Packing",                5, "Frontend Dev", "5", "Navneet", "Per-product transit packaging + monthly volume + cost calc", "Completed", "UAT-1", D(2026,3,12),  D(2026,3,16),  D(2026,3,12),  D(2026,3,16)),
    (2, "6.24", "Logistics - Service",                5, "Frontend Dev", "5", "Navneet", "Connectivity + available services + route blocks with leg tables", "Completed", "UAT-1", D(2026,3,17),  D(2026,3,21),  D(2026,3,17),  D(2026,3,21)),
    (2, "6.25", "Infrastructure (BE + EE + CE combined)", 67, "Frontend Dev", "5", "Navneet", "BE plinth rate + EE auto from machinery + CE solarisation + Financial Model (PRIME/Innovation)", "In Progress", "UAT-1", D(2026,3,22),  D(2026,5,28),  D(2026,3,22), None),
    (2, "6.26", "Funding - Pitch Deck",               3, "Frontend Dev", "5", "Navneet", "Summary (X/11 progress) + Existing + Required (11 slides) + log book", "Completed", "UAT-1", D(2026,3,23),  D(2026,3,25),  D(2026,3,23),  D(2026,3,25)),
    (2, "6.27", "Funding - Schemes",                  5, "Frontend Dev", "5", "Navneet", "17-scheme master + loan details + EMI calc + Required pipeline", "Completed", "UAT-1", D(2026,3,26),  D(2026,3,30),  D(2026,3,26),  D(2026,3,30)),

    # WBS 7 -- Research Notes (UAT-2)
    (1, "7", "Research Notes", 6, "Developer", "3", "Navneet",
     "Fessociates can log free-form research notes; optional EP linking",
     "In Progress", "UAT-2", D(2026,3,18), D(2026,6,5), D(2026,3,18), None),
    (2, "7.1", "Research Notes doctype backend", 2, "Backend Dev", "3", "Navneet",
     "Doctype with 11 fields deployed; permissions wired",
     "Completed", "UAT-2", D(2026,3,18), D(2026,3,19), D(2026,3,18), D(2026,3,19)),
    (2, "7.2", "Vue: Research Notes form (11 fields)", 4, "Frontend Dev", "7.1", "Navneet",
     "Standard Frappe form with 11 fields renders + saves",
     "Completed", "UAT-2", D(2026,3,20), D(2026,3,23), D(2026,3,20), D(2026,3,23)),
    (2, "7.3", "EP Profile linking (optional)", 5, "Backend Dev", "7.1,4.1", "Navneet",
     "Note can be tagged to an entrepreneur; surfaces on EP profile",
     "Not Started", "UAT-2", D(2026,6,2), D(2026,6,5), None, None),

    # WBS 8 -- Village Survey (UAT-2)
    (1, "8", "Village Survey", 12, "Developer", "3", "Navneet",
     "Village-level survey form complete; geo-pins ready for Landscape Map",
     "In Progress", "UAT-2", D(2026,3,24), D(2026,6,9), D(2026,3,24), None),
    (2, "8.1", "Village Survey doctype backend", 2, "Backend Dev", "3", "Navneet",
     "village_visit doctype with 16 picker child doctypes deployed",
     "Completed", "UAT-2", D(2026,3,24), D(2026,3,25), D(2026,3,24), D(2026,3,25)),
    (2, "8.2", "Vue: Survey form (17 fields + geolocation)", 5, "Frontend Dev", "8.1", "Navneet",
     "Form renders + saves; geolocation captured",
     "Completed", "UAT-2", D(2026,3,26), D(2026,3,30), D(2026,3,26), D(2026,3,30)),
    (2, "8.3", "Master pickers: crops + forest resources + livelihoods (16 pickers)", 5, "Backend Dev", "8.1", "Navneet",
     "All 16 picker doctypes loaded with seed data",
     "Completed", "UAT-2", D(2026,3,31), D(2026,4,4), D(2026,3,31), D(2026,4,4)),
    (2, "8.4", "Geo-tag -> Landscape Map integration (village pins)", 6, "Backend Dev", "8.2,10.1", "Navneet",
     "Village pins render on Landscape Map with popup",
     "Not Started", "UAT-2", D(2026,6,2), D(2026,6,9), None, None),

    # WBS 9 -- Fellow Activity Logger (UAT-1)
    (1, "9", "Fellow Activity Logger", 60, "Developer", "5,6", "Navneet",
     "Fessociates log activities (Calendar/Board) with cascading filters; sync to EP + Log Book",
     "In Progress", "UAT-1", D(2026,4,5), D(2026,5,28), D(2026,4,5), None),
    (2, "9.1", "Replace Tasks workspace with Activity Logger pages", 4, "Frontend Dev", "5,6", "Navneet",
     "task_manager + task_manager_app + task_manager_grid pages deployed",
     "Completed", "UAT-1", D(2026,4,5), D(2026,4,8), D(2026,4,5), D(2026,4,8)),
    (2, "9.2", "Vue: Calendar view (default)", 7, "Frontend Dev", "9.1", "Navneet",
     "TaskManagerApp.vue renders calendar with activities",
     "Completed", "UAT-1", D(2026,4,9), D(2026,4,15), D(2026,4,9), D(2026,4,15)),
    (2, "9.3", "Vue: Board view + toggle", 4, "Frontend Dev", "9.1", "Navneet",
     "TaskManagerGrid.vue renders board; toggle between board/calendar",
     "Completed", "UAT-1", D(2026,4,16), D(2026,4,19), D(2026,4,16), D(2026,4,19)),
    (2, "9.4", "6 cascading filters (District -> Block -> EP -> Module -> Framework -> Task)", 38, "Frontend Dev", "9.2", "Navneet",
     "All 6 levels cascade and reset downstream selections",
     "In Progress", "UAT-1", D(2026,4,20), D(2026,5,28), D(2026,4,20), None),
    (2, "9.5", "Activity form: branching logic (3 paths: EP-Framework / EP-NonFramework / NonEntrepreneur)", 35, "Frontend Dev", "9.2", "Navneet",
     "Activity form branches correctly based on entrepreneur + framework selection",
     "In Progress", "UAT-1", D(2026,4,24), D(2026,5,28), D(2026,4,24), None),
    (2, "9.6", "Sync: framework activities -> Log Book tabs", 5, "Backend Dev", "9.5", "Navneet",
     "Activity completion writes to corresponding framework Log Book",
     "Not Started", "UAT-2", D(2026,6,2), D(2026,6,6), None, None),
    (2, "9.7", "Sync: all EP activities -> Intervention Logs (read-only)", 5, "Backend Dev", "9.5,4.8", "Navneet",
     "EP Profile -> Intervention Logs reflects every Activity Logger entry",
     "Not Started", "UAT-2", D(2026,6,2), D(2026,6,6), None, None),
    (2, "9.8", "Framework completion tracking (Not Started -> Completed -> xN)", 7, "Backend Dev", "9.7", "Navneet",
     "Each framework status auto-updates on EP dashboard per activity completion",
     "Not Started", "UAT-2", D(2026,6,3), D(2026,6,9), None, None),

    # WBS 10 -- Landscape Dashboard (UAT-2)
    (1, "10", "Landscape Dashboard", 60, "Developer", "4,8", "Navneet",
     "Entrepreneurial Landscape Map with 6 filters + clustering + popup + village pins",
     "In Progress", "UAT-2", D(2026,4,10), D(2026,6,9), D(2026,4,10), None),
    (2, "10.1", "Vue: Entrepreneurial Landscape Map (6 filters + clustering + popup)", 50, "Frontend Dev", "4,8", "Navneet",
     "Map renders all EPs as pins; 6-filter narrowing live; cluster on zoom-out",
     "In Progress", "UAT-2", D(2026,4,10), D(2026,6,5), D(2026,4,10), None),
    (2, "10.2", "Village Survey geo-pin overlay on Landscape Map", 8, "Backend Dev", "8.4,10.1", "Navneet",
     "Village pins shown alongside EP pins with distinct icon",
     "Not Started", "UAT-2", D(2026,6,2), D(2026,6,9), None, None),

    # WBS 11 -- Milestone & Activity Tracker (UAT-2 - already live)
    (1, "11", "Milestone & Activity Tracker", 12, "Developer / Tech Lead", "1", "Navneet",
     "Progress Matrix workspace with Overview/Activities/Timeline/RACI tabs [Bug 55 P1: dashboard non-functional, REGRESSED in UAT-1]",
     "In Progress", "UAT-2", D(2026,4,1), D(2026,4,12), D(2026,4,1), None),
    (2, "11.1", "Vue: Milestone Tracker dashboard (4 tabs + lollipop + Gantt)", 8, "Frontend Dev", "1", "Navneet",
     "UI built; but Bug 55 (P1): get_milestone_data ModuleNotFoundError -> 'Error loading data' + 0 records for ALL roles",
     "In Progress", "UAT-2", D(2026,4,1), D(2026,4,8), D(2026,4,1), None),
    (2, "11.2", "Workspace + Web Page + CHB deployment (Progress Matrix)", 1, "Tech Lead", "11.1", "Bhushan",
     "/prime-rural-milestone-tracker live; Progress Matrix workspace seeded",
     "Completed", "UAT-2", D(2026,4,9), D(2026,4,9), D(2026,4,9), D(2026,4,9)),
    (2, "11.3", "Milestone Master + Milestone Activity records populated (26 + 85)", 3, "PM", "11.2", "Samarth",
     "26 milestones loaded; the 85 activities were lost in the Bug-55 regression - needs reseed",
     "In Progress", "UAT-2", D(2026,4,10), D(2026,4,12), D(2026,4,10), None),

    # WBS 12 -- Intervention Dashboard & KPI Form (UAT-3, Pending Requirements)
    (1, "12", "Intervention Dashboard & KPI Form", None, "Developer + PM + Client", "", "Navneet",
     "Requirement finalisation needed - KPI form spec pending from client (post 20-May meeting)",
     "Pending Requirements", "UAT-3", None, None, None, None),
    (2, "12.1", "KPI form spec gathering from client (Minha + Alem)", None, "PM + Client", "", "Samarth",
     "KPI list + form fields + dashboard wireframe signed off by MBMA",
     "Pending Requirements", "UAT-3", None, None, None, None),
    (2, "12.2", "Intervention Dashboard rebuild (KPIs + Sankey + drill-down)", None, "Frontend Dev", "12.1", "Navneet",
     "Dashboard reflects signed-off KPIs with live data + drill-down",
     "Pending Requirements", "UAT-3", None, None, None, None),
    (2, "12.3", "Sankey Flow component (already built; needs KPI wiring)", 7, "Frontend Dev", "", "Navneet",
     "SankeyFlow.vue renders MET -> Incubated/Supported -> sectors",
     "Completed", "Pre-UAT", D(2026,4,11), D(2026,4,17), D(2026,4,11), D(2026,4,17)),

    # WBS 13 -- Apoorva AI Integration (UAT-3, Pending Requirements)
    (1, "13", "Apoorva AI Integration", None, "Developer + Client", "", "Navneet",
     "Requirement finalisation needed - integration spec pending from Apoorva AI team",
     "Pending Requirements", "UAT-3", None, None, None, None),
    (2, "13.1", "Integration spec with Apoorva AI team (endpoints + auth + context model)", None, "PM + Client", "", "Samarth",
     "API contract signed; auth + rate limits agreed",
     "Pending Requirements", "UAT-3", None, None, None, None),
    (2, "13.2", "Chatbot widget on portal (entrepreneur-facing)", None, "Frontend Dev", "13.1", "Navneet",
     "Widget docks on portal; sends + receives messages",
     "Pending Requirements", "UAT-3", None, None, None, None),
    (2, "13.3", "AI Generate: Label Storyline (in Labels Required tab)", None, "Backend Dev", "13.1", "Navneet",
     "Button on Label form generates storyline draft via Apoorva",
     "Pending Requirements", "UAT-3", None, None, None, None),
    (2, "13.4", "AI Generate: Brochure Storyline (in Brochure form)", None, "Backend Dev", "13.1", "Navneet",
     "Button on Brochure form generates storyline draft via Apoorva",
     "Pending Requirements", "UAT-3", None, None, None, None),
    (2, "13.5", "Context-aware responses (pass entrepreneur data to model)", None, "Backend Dev", "13.1", "Navneet",
     "Chatbot answers reference the current entrepreneur's data",
     "Pending Requirements", "UAT-3", None, None, None, None),

    # WBS 14 -- Claude-based Chatbot (UAT-3, Pending Requirements)
    (1, "14", "Claude-based Chatbot", None, "Developer + Client", "", "Navneet",
     "Requirement finalisation needed - scope of Claude vs Apoorva to be agreed",
     "Pending Requirements", "UAT-3", None, None, None, None),
    (2, "14.1", "Claude integration spec (MCP scope, tool list, auth model)", None, "PM + Client", "", "Samarth",
     "Tool list + auth model + which user roles get access agreed",
     "Pending Requirements", "UAT-3", None, None, None, None),
    (2, "14.2", "Internal-team chatbot widget (Core Team IT / PM use)", None, "Frontend Dev", "14.1", "Navneet",
     "Chatbot panel for internal users with MCP-powered queries",
     "Pending Requirements", "UAT-3", None, None, None, None),
    (2, "14.3", "Tool integration: entrepreneur queries, status updates, dashboard pulls", None, "Backend Dev", "14.1", "Navneet",
     "MCP server exposes write tools beyond current read-only set",
     "Pending Requirements", "UAT-3", None, None, None, None),

    # WBS 15 -- QA, VAPT & Performance
    (1, "15", "QA, VAPT & Performance", 40, "Testing / DevOps / Tech Lead", "all", "Chitranshi",
     "All UAT cycles run; VAPT + load tests cleared before go-live",
     "In Progress", "UAT-1", D(2026,5,25), D(2026,7,3), D(2026,5,25), None),
    (2, "15.1", "UAT environment setup (separate from staging)", 3, "DevOps", "1", "Prakash",
     "uat.prime-rural.dhwaniris.in live with prod-like data snapshot",
     "In Progress", "UAT-1", D(2026,5,25), D(2026,5,28), D(2026,5,25), None),
    (2, "15.2", "UAT-1 test scripts (User Mgmt + Masters + EP + Product + 27 frameworks + Activity Logger)", 3, "QA Engineer", "15.1", "Chitranshi",
     "330 TDD cases documented + Playwright E2E suite committed; checklist live on GitHub Pages",
     "Completed", "UAT-1", D(2026,5,26), D(2026,6,7), D(2026,5,26), D(2026,6,7)),
    (2, "15.3", "UAT-1 internal dry run (E2E against staging)", 1, "QA + PM", "15.2", "Chitranshi / Samarth",
     "E2E executed 7-8 Jun: 272 cases run (67 pass / 37 fail / 168 blocked), ~60 bugs logged; 3 P1 fixes pending",
     "In Progress", "UAT-1", D(2026,5,29), D(2026,6,8), D(2026,5,29), None),
    (2, "15.4", "UAT-1 with client (MBMA team)", 1, "PM + Client", "15.3", "Samarth",
     "Client signs off UAT-1 scope or files defects",
     "Not Started", "UAT-1", D(2026,5,29), D(2026,5,29), None, None),
    (2, "15.5", "UAT-1 bug fixing sprint", 5, "Tech Lead + Developer", "15.4", "Bhushan",
     "All UAT-1 P0/P1 defects fixed + redeployed",
     "Not Started", "UAT-1", D(2026,5,30), D(2026,6,5), None, None),
    (2, "15.6", "UAT-2 test scripts (Research Notes + Village Survey + Landscape + Milestone Tracker)", 4, "QA Engineer", "15.5", "Chitranshi",
     "Test scripts for UAT-2 scope signed off", "Not Started", "UAT-2",
     D(2026,6,8), D(2026,6,11), None, None),
    (2, "15.7", "UAT-2 with client (MBMA team)", 1, "PM + Client", "15.6", "Samarth",
     "Client signs off UAT-2 scope", "Not Started", "UAT-2",
     D(2026,6,12), D(2026,6,12), None, None),
    (2, "15.8", "UAT-2 bug fixing sprint", 5, "Tech Lead + Developer", "15.7", "Bhushan",
     "All UAT-2 P0/P1 defects fixed + redeployed", "Not Started", "UAT-2",
     D(2026,6,13), D(2026,6,19), None, None),
    (2, "15.9", "OWASP ZAP scan + role-based access + SQLi/XSS testing", 5, "Security Eng", "15.5", "Chitranshi",
     "VAPT report - no Critical/High open findings",
     "Not Started", "Pre-Go-Live", D(2026,6,22), D(2026,6,26), None, None),
    (2, "15.10", "Load test: 80 concurrent + 1200 EPs + 2000 map pins", 5, "QA Engineer", "15.5", "Chitranshi",
     "Page load < 3s; API < 1s; map renders smoothly",
     "Not Started", "Pre-Go-Live", D(2026,6,29), D(2026,7,3), None, None),

    # WBS 16 -- UAT & Go-Live
    (1, "16", "UAT & Go-Live", 25, "Developer / DevOps / PM / Client", "15", "Samarth",
     "Production cutover with MBMA sign-off + 80 fellows onboarded",
     "Not Started", "Go-Live", D(2026,7,6), D(2026,8,3), None, None),
    (2, "16.1", "UAT-3 scope finalization (post requirement gathering)", None, "PM + Client", "12.1,13.1,14.1", "Samarth",
     "UAT-3 scope: Intervention Dashboard + Apoorva AI + Claude Chatbot - dates set once specs sign off",
     "Pending Requirements", "UAT-3", None, None, None, None),
    (2, "16.2", "Production environment setup + SSL + domain", 5, "DevOps", "15", "Prakash",
     "prime-rural.dhwaniris.in live with HTTPS + backups configured",
     "Not Started", "Go-Live", D(2026,7,6), D(2026,7,10), None, None),
    (2, "16.3", "Data migration: Google Sheets -> Frappe", 5, "Backend Dev", "16.2", "Navneet",
     "Legacy 32-framework data + entrepreneur records migrated + validated",
     "Not Started", "Go-Live", D(2026,7,13), D(2026,7,17), None, None),
    (2, "16.4", "User onboarding: 80 fellows + coordinators", 5, "PM", "16.3", "Samarth",
     "All users invited; passwords set; first login captured",
     "Not Started", "Go-Live", D(2026,7,20), D(2026,7,24), None, None),
    (2, "16.5", "Training sessions for field users (Fessociates + Coordinators)", 5, "PM", "16.4", "Samarth",
     "All 80 users trained; recordings shared; FAQ doc published",
     "Not Started", "Go-Live", D(2026,7,27), D(2026,7,31), None, None),
    (2, "16.6", "Go-live sign-off from MBMA", 1, "PM + Client", "16.5", "Samarth",
     "Signed handover note + MoM filed",
     "Not Started", "Go-Live", D(2026,8,3), D(2026,8,3), None, None),
]

# -------- Status -> milestone row color --------
STATUS_FILL = {
    "Completed":            "C8E6C9",  # light green
    "In Progress":          "BBDEFB",  # light blue
    "Pending Requirements": "FFE0B2",  # light orange
    "Not Started":          "F5F5F5",  # light grey
}
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1A6B3C")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
MILESTONE_FONT = Font(bold=True, size=11)
SUB_FONT = Font(size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# -------- Build workbook --------
wb = Workbook()
del wb["Sheet"]

# ===== Sheet: Header =====
ws = wb.create_sheet("Header")
header_lines = [
    ("DHWANI RIS - PROJECT PLAN", "", ""),
    ("", "", ""),
    ("Project Name", "PRIME Rural Fellowship Platform - Web CRM", ""),
    ("Project ID", "PREF-CRM-WEB", ""),
    ("Client", "MBMA (Meghalaya Basin Management Agency)", ""),
    ("Delivery Team", "Dhwani Rural Information Services", ""),
    ("", "", ""),
    ("Project Start Date", D(2025,9,1), ""),
    ("Planned Go-Live Date", D(2026,8,3), ""),
    ("Plan Version", "v6.1", ""),
    ("Plan Date", D(2026,6,9), ""),
    ("Plan Author", "Samarth Paul (PM)", ""),
    ("", "", ""),
    ("UAT-1 Date", D(2026,5,29), "User Mgmt + Masters + EP Profile + Product Profile + 27 Frameworks + Fellow Activity Logger"),
    ("UAT-2 Date (proposed)", D(2026,6,12), "Research Notes + Village Survey + Landscape Dashboard + Milestone Tracker polish"),
    ("UAT-3 Date", "TBD", "Intervention Dashboard + KPI + Apoorva AI + Claude Chatbot - pending requirement finalisation"),
]
for r, row in enumerate(header_lines, 1):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        if r == 1:
            cell.font = Font(bold=True, size=14, color="1A6B3C")
        elif c == 1 and val:
            cell.font = Font(bold=True)
        if isinstance(val, date):
            cell.number_format = "DD-MMM-YYYY"
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 70

# ===== Sheet: Project Plan =====
ws = wb.create_sheet("Project Plan")

# Header row
for c, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER
ws.row_dimensions[1].height = 36

for i, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Data rows
sno = 1
for row in ROWS:
    level, wbs, activity, dur, res, dep, owner, succ, status, batch, sd, ed, asd, aed = row
    values = [sno, wbs, activity, dur, res, dep, owner, succ, status, batch, sd, ed, asd, aed]
    excel_row = ws.max_row + 1
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=excel_row, column=c, value=v)
        cell.border = BORDER
        cell.alignment = WRAP if c in (3, 5, 8) else Alignment(vertical="top", wrap_text=True)
        if isinstance(v, date):
            cell.number_format = "DD-MMM-YYYY"
        if level == 1:
            cell.fill = PatternFill("solid", fgColor=STATUS_FILL.get(status, "F5F5F5"))
            cell.font = MILESTONE_FONT
        else:
            cell.font = SUB_FONT
    sno += 1

# Freeze header + first 2 cols for navigation
ws.freeze_panes = "C2"

# AutoFilter on full data range
last_col_letter = get_column_letter(len(HEADERS))
ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

# ===== Sheet: Gantt Chart (summary roll-up) =====
ws = wb.create_sheet("Gantt Chart")
ws.append(["UAT Phase", "Milestone (WBS L1)", "Status", "Planned Start", "Planned End", "Notes"])
for c in range(1, 7):
    cell = ws.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER

# Build summary from milestone rows
milestone_rows = [r for r in ROWS if r[0] == 1]
for r in milestone_rows:
    _, wbs, activity, _, _, _, _, _, status, batch, sd, ed, _, _ = r
    note = ""
    if status == "Pending Requirements":
        note = "Dates TBD - requirement finalisation needed"
    elif batch == "UAT-1":
        note = "Target UAT-1: 29-May-2026"
    elif batch == "UAT-2":
        note = "Target UAT-2: 12-Jun-2026 (proposed)"
    elif batch == "UAT-3":
        note = "Target UAT-3: TBD (post-requirements)"
    excel_row = ws.max_row + 1
    values = [batch, f"{wbs}. {activity}", status, sd, ed, note]
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=excel_row, column=c, value=v)
        cell.border = BORDER
        cell.alignment = WRAP
        if isinstance(v, date):
            cell.number_format = "DD-MMM-YYYY"
        if c == 3:  # status cell
            cell.fill = PatternFill("solid", fgColor=STATUS_FILL.get(status, "F5F5F5"))
            cell.font = Font(bold=True)

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 50
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:F{ws.max_row}"

# ===== Sheet: How to use this document =====
ws = wb.create_sheet("How to use this document")
howto = [
    ("Before you begin", ""),
    ("1", "Read the Dhwani Setu methodology PM template guidance before editing this plan."),
    ("2", "Save a backup before bulk-editing rows. Update Plan Version + Plan Date on Header sheet on every change."),
    ("", ""),
    ("How to use this document", ""),
    ("1", "Project Plan sheet is the master. WBS L1 rows are milestones (colored by status). L2 rows are sub-activities (white)."),
    ("2", "UAT Batch column groups activities by delivery phase: UAT-1 (29-May-2026), UAT-2 (12-Jun-2026 proposed), UAT-3 (TBD post-requirements), Pre-UAT, Pre-Go-Live, Go-Live."),
    ("3", "AutoFilter is on - use the dropdown arrows on row 1 to filter by Status, Owner, UAT Batch, etc."),
    ("4", "Status values: Completed (green), In Progress (blue), Pending Requirements (orange), Not Started (grey). Pending Requirements means spec from client/SME is needed before dates can be set."),
    ("5", "Owners follow the project mapping: Navneet (developer), Bhushan (tech lead), Samarth (PM), Chitranshi/Samarth (testing), Prakash (DevOps), Client (MBMA / Apoorva / SMEs)."),
    ("6", "Dates: Start/End = planned. Actual Start/End populated on completion. For Completed items, Planned and Actual match (backdated from real completion)."),
    ("7", "Gantt Chart sheet is a milestone-level roll-up grouped by UAT Batch."),
    ("8", "Cells are never merged in the Project Plan data area, so filtering and sorting always work."),
    ("", ""),
    ("Change log", ""),
    ("-", "v6 rebuilt from v5 CSV and the actual repo state. Added UAT Batch column, owner reassignment, and backdated planned dates."),
]
for r, (a, b) in enumerate(howto, 1):
    ws.cell(row=r, column=1, value=a).font = Font(bold=True) if a and not a.isdigit() else Font()
    ws.cell(row=r, column=2, value=b).alignment = WRAP
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 110

# ===== Sheet: Version Control =====
ws = wb.create_sheet("Version Control")
vc_hdr = ["S No", "Version", "Date", "Changes", "Made By", "Approved By", "Current State"]
for c, h in enumerate(vc_hdr, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
prior = [
    [1, "v3", D(2026,3,23), "Initial client xlsx (329 rows). Columns: WBS, Activity, Duration, Resources, Dependency, Owner, Success Criteria, Current Status, Start/End/Actual dates.", "Dhwani PM team", "MBMA", "Superseded"],
    [2, "v4", D(2026,5,10), "Mid-cycle refresh. Statuses adjusted manually. Not repo-validated.", "Dhwani PM team", "-", "Superseded"],
    [3, "v5", D(2026,5,15), "CSV rebuild against repo (234 rows). Stripped date/duration/success criteria columns. Used to populate generated/data/project-plan.js -> plan.html.", "Samarth (PM) + Claude", "-", "Superseded"],
    [4, "v6", D(2026,5,26), "Cut on Dhwani Setu PM template. 16 WBS roots: 14 user-named milestones + System Foundation + UAT/Go-Live. Owners reassigned (Navneet=Dev, Bhushan=TL, Samarth=PM, Chitranshi=Test, Prakash=DevOps, Client). UAT Batch column added. Planned dates backdated from actuals; tight clustering reflects AI-assisted build velocity. Intervention Dashboard, Apoorva AI, Claude Chatbot left at 'Pending Requirements' with no dates.", "Samarth (PM) + Claude", "Pending MBMA review", "Superseded"],
    [5, "v6.1", D(2026,6,9), "Status reconciliation against the 7-8 Jun internal UAT-1 E2E run (272 cases, ~60 bugs) + the Jun 1-3 dev sprint. User Management (2.1/2.2) -> Completed (verified). UAT-1 test scripts (15.2) -> Completed; internal dry run (15.3) -> In Progress (executed, 3 P1 fixes pending). Prototyping (6.10) and Milestone Tracker (11/11.1/11.3) -> In Progress after P1 regressions (Bug 54 save-500, Bug 55 dashboard non-functional). No new rows.", "Samarth (PM) + Claude", "Pending MBMA review", "Published"],
]
for r, row in enumerate(prior, 2):
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.alignment = WRAP
        if isinstance(v, date):
            cell.number_format = "DD-MMM-YYYY"
for col, width in zip("ABCDEFG", [6, 9, 13, 75, 22, 22, 18]):
    ws.column_dimensions[col].width = width

# Save
wb.save(OUT)
print("Wrote:", OUT)
print("Rows in Project Plan:", len([r for r in ROWS]))
print("Milestones (WBS L1):", len([r for r in ROWS if r[0] == 1]))
print("Sub-activities (WBS L2):", len([r for r in ROWS if r[0] == 2]))

# -------- Also emit data/project-plan.js for plan.html --------
import json, os

JS_OUT = r"C:\Users\Dhwani\OneDrive - Dhwani Rural Information Systems Pvt. Ltd\Work\prime-rural\generated\data\project-plan.js"

def fmt_date(d):
    return d.strftime("%d-%b-%Y") if isinstance(d, date) else ""

records = []
current_module = ""
sno = 1
for row in ROWS:
    level, wbs, activity, dur, res, dep, owner, succ, status, batch, sd, ed, asd, aed = row
    if level == 1:
        current_module = activity  # milestone becomes the module/group label
    records.append({
        "sno": sno,
        "wbs": wbs,
        "activity": activity,
        "level": level,
        "module": current_module,
        "duration": dur if dur is not None else "",
        "resources": res or "",
        "dependency": dep or "",
        "owner": owner or "",
        "successCriteria": succ or "",
        "status": status or "",
        "uatBatch": batch or "",
        "startDate": fmt_date(sd),
        "endDate": fmt_date(ed),
        "actualStart": fmt_date(asd),
        "actualEnd": fmt_date(aed),
    })
    sno += 1

os.makedirs(os.path.dirname(JS_OUT), exist_ok=True)
with open(JS_OUT, "w", encoding="utf-8") as f:
    f.write("/* Auto-generated from PREF_CRM_Web_Project_Plan_v6 source-of-truth (build script: generated/_build_v6_plan.py). */\n")
    f.write("var PROJECT_PLAN = ")
    json.dump(records, f, indent=1, ensure_ascii=False)
    f.write(";\n")
print("Wrote:", JS_OUT, "(", len(records), "records )")
